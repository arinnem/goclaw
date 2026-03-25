"""
v4 Part 3: Revenue + Cost + P&L + Cash Flow + DCF
"""
import openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/v4_state.json') as f: state=json.load(f)
OUT=state['out']

HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LBF=Font(size=11);TF=Font(bold=True,size=11)
TFL=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FMT='#,##0';FP='0.0%'

def hd(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def sc(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
def lb(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LBF;cl.border=BD
def fm(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=Font(size=11);cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def tot(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=TF;cl.fill=TFL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)

MN=['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
TC=[]
for m in range(7,13):TC.append((f"{MN[m]} 2025",2025,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2026",2026,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2027",2027,'m',m))
for yr in [2028,2029,2030]:
    for q in range(1,5):TC.append((f"Q{q} {yr}",yr,'q',q))
TSC=3;NTC=42

def time_hdrs(ws,r):
    hd(ws,r,1,"Item");hd(ws,r,2,"Unit")
    for i,(l,y,t,m) in enumerate(TC):col=TSC+i;hd(ws,r,col,l);ws.column_dimensions[get_column_letter(col)].width=11

net_r=state['net_r'];vcr_r=state['vcr_r'];ecom_r=state['ecom_r']
div_r=state['div_r'];mon_r=state['mon_r'];ybase_r=state['ybase_r'];yfrac_r=state['yfrac_r']
hr_tot_r=state['hr_tot_r']
AN=[(0,5,'H2_2025'),(6,17,'Y2026'),(18,29,'Y2027'),(30,33,'Y2028'),(34,37,'Y2029'),(38,41,'Y2030')]

def ms(row,i):return f"'Market Sizing'!{get_column_letter(TSC+i)}{row}"

wb=openpyxl.load_workbook(OUT)

# ===== REVENUE =====
ws=wb.create_sheet("Revenue Schedule");ws.sheet_properties.tabColor="00B050"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"REVENUE SCHEDULE (M VND)");r+=2;time_hdrs(ws,r);r+=1

sc(ws,r,1,"STAGE 1");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1

cert_r=r
lb(ws,r,1,"  Certification fees");lb(ws,r,2,"M VND")
for i in range(NTC):
    fm(ws,r,TSC+i,f"={ms(net_r,i)}*cert_fee_yr1*(1+cert_fee_escalation)^{ms(ybase_r,i)}/{ms(div_r,i)}",FMT)
r+=1
train_r=r
lb(ws,r,1,"  Training");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={ms(net_r,i)}*training_pct*training_fee/{ms(div_r,i)}",FMT)
r+=1
audit_r=r
lb(ws,r,1,"  Audit fees");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={ms(net_r,i)}*audit_fee/{ms(div_r,i)}",FMT)
r+=1
s1_r=r
lb(ws,r,1,"Subtotal Stage 1")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"=SUM({cl}{cert_r}:{cl}{audit_r})",FMT)
r+=2

sc(ws,r,1,"STAGE 2 (gated)")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
tx_r=r
lb(ws,r,1,"  Transaction fees");lb(ws,r,2,"M VND")
for i in range(NTC):
    fm(ws,r,TSC+i,f"=IF({ms(net_r,i)}>=stage2_gate,{ms(ecom_r,i)}*platform_market_share*tx_take_rate*1000,0)",FMT)
r+=1
daas_r=r
lb(ws,r,1,"  DaaS");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=IF({ms(net_r,i)}>=stage2_gate,{ms(vcr_r,i)}*daas_price,0)",FMT)
r+=1
ad_r=r
lb(ws,r,1,"  Advertising");lb(ws,r,2,"M VND")
for i in range(NTC):
    fm(ws,r,TSC+i,f"=IF({ms(net_r,i)}>=stage2_gate,ad_impressions_base*{ms(mon_r,i)}*{ms(net_r,i)}/1000*ad_cpm/1000,0)",FMT)
r+=1
s2_r=r
lb(ws,r,1,"Subtotal Stage 2")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"=SUM({cl}{tx_r}:{cl}{ad_r})",FMT)
r+=2

rev_tot_r=r
lb(ws,r,1,"TOTAL REVENUE")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{s1_r}+{cl}{s2_r}",FMT)
r+=2

sc(ws,r,1,"COGS")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
cogs_r=r
lb(ws,r,1,"  Payment processing");lb(ws,r,2,"M VND")
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"={cl}{tx_r}*payment_proc_pct",FMT)
r+=1
cogs_tot_r=r
lb(ws,r,1,"TOTAL COGS")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{cogs_r}",FMT)

# ===== COST SCHEDULE =====
ws=wb.create_sheet("Cost Schedule");ws.sheet_properties.tabColor="FF6600"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"COST SCHEDULE (M VND)");r+=2;time_hdrs(ws,r);r+=1

sc(ws,r,1,"CAPEX")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
cdev_r=r
lb(ws,r,1,"  Platform dev");lb(ws,r,2,"M VND")
for i,(l,yr,t,m) in enumerate(TC):
    if i<6:fm(ws,r,TSC+i,"=capex_total/6",FMT)
    else:fm(ws,r,TSC+i,"=0",FMT)
r+=1
cmnt_r=r
lb(ws,r,1,"  Maintenance");lb(ws,r,2,"M VND")
for i,(l,yr,t,m) in enumerate(TC):
    if yr<=2025:fm(ws,r,TSC+i,"=0",FMT)
    else:fm(ws,r,TSC+i,f"=capex_total*capex_maintenance_pct/{ms(div_r,i)}",FMT)
r+=1
cx_tot_r=r
lb(ws,r,1,"TOTAL CAPEX")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{cdev_r}+{cl}{cmnt_r}",FMT)
r+=2

sc(ws,r,1,"OPEX")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1

def hr_c(row,i):return f"'HR Plan'!{get_column_letter(TSC+i)}{row}"

opex_start=r
lb(ws,r,1,"  HR (from HR Plan)");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={hr_c(hr_tot_r,i)}",FMT)
r+=1
lb(ws,r,1,"  Marketing");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=opex_marketing*(1+marketing_growth)^{ms(ybase_r,i)}/{ms(div_r,i)}",FMT)
r+=1
lb(ws,r,1,"  Branding & events");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=opex_branding/{ms(div_r,i)}",FMT)
r+=1
lb(ws,r,1,"  Cloud & infra");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=opex_cloud_base/{ms(div_r,i)}+{ms(vcr_r,i)}*cloud_per_1M_vcr",FMT)
r+=1
lb(ws,r,1,"  Security");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=opex_security/{ms(div_r,i)}",FMT)
r+=1
lb(ws,r,1,"  External affairs");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=opex_external/{ms(div_r,i)}",FMT)
r+=1
lb(ws,r,1,"  Legal & compliance");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=opex_legal/{ms(div_r,i)}",FMT)
r+=1
pilot_r=r
lb(ws,r,1,"  Pilot eval (Y1)");lb(ws,r,2,"M VND")
for i,(l,yr,t,m) in enumerate(TC):
    if yr==2025:fm(ws,r,TSC+i,f"=opex_pilot_eval/{ms(div_r,i)}",FMT)
    else:fm(ws,r,TSC+i,"=0",FMT)
r+=1
lb(ws,r,1,"  Partner mgmt");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=opex_partner_mgmt/{ms(div_r,i)}",FMT)
r+=1
lb(ws,r,1,"  Customer support");lb(ws,r,2,"M VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={ms(net_r,i)}*opex_support_per_ent/{ms(div_r,i)}",FMT)
opex_end=r;r+=1

ox_sub_r=r
lb(ws,r,1,"OPEX Subtotal")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"=SUM({cl}{opex_start}:{cl}{opex_end})",FMT)
r+=1
cont_r=r
lb(ws,r,1,"  Contingency");lb(ws,r,2,"M VND")
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"={cl}{ox_sub_r}*opex_contingency_pct",FMT)
r+=1
ox_tot_r=r
lb(ws,r,1,"TOTAL OPEX")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ox_sub_r}+{cl}{cont_r}",FMT)
r+=2
tot_cost_r=r
lb(ws,r,1,"TOTAL COSTS")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{cx_tot_r}+{cl}{ox_tot_r}",FMT)

# ===== P&L =====
ws=wb.create_sheet("P&L");ws.sheet_properties.tabColor="7030A0"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"P&L (M VND)");r+=2;time_hdrs(ws,r);r+=1

def rv(row,i):return f"'Revenue Schedule'!{get_column_letter(TSC+i)}{row}"
def cs(row,i):return f"'Cost Schedule'!{get_column_letter(TSC+i)}{row}"

rev_pnl=r;lb(ws,r,1,"Revenue")
for i in range(NTC):fm(ws,r,TSC+i,f"={rv(rev_tot_r,i)}",FMT)
r+=1
cogs_pnl=r;lb(ws,r,1,"COGS")
for i in range(NTC):fm(ws,r,TSC+i,f"=-{rv(cogs_tot_r,i)}",FMT)
r+=1
gp_r=r;lb(ws,r,1,"GROSS PROFIT")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{rev_pnl}+{cl}{cogs_pnl}",FMT)
r+=2
opex_pnl=r;lb(ws,r,1,"OPEX")
for i in range(NTC):fm(ws,r,TSC+i,f"=-{cs(ox_tot_r,i)}",FMT)
r+=1
ebitda_r=r;lb(ws,r,1,"EBITDA")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{gp_r}+{cl}{opex_pnl}",FMT)
r+=1
dep_pnl=r;lb(ws,r,1,"Depreciation")
for i,(l,yr,t,m) in enumerate(TC):
    if yr<=2029:fm(ws,r,TSC+i,f"=-capex_total/depreciation_years/{ms(div_r,i)}",FMT)
    else:fm(ws,r,TSC+i,"=0",FMT)
r+=1
ebit_r=r;lb(ws,r,1,"EBIT")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ebitda_r}+{cl}{dep_pnl}",FMT)
r+=1
tax_r=r;lb(ws,r,1,"Tax (CIT)")
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"=IF({cl}{ebit_r}>0,-{cl}{ebit_r}*cit_rate,0)",FMT)
r+=1
ni_r=r;lb(ws,r,1,"NET INCOME")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ebit_r}+{cl}{tax_r}",FMT)
r+=2
# Margins
sc(ws,r,1,"MARGINS");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
for lab,nr in[("Gross margin",gp_r),("EBITDA margin",ebitda_r),("Net margin",ni_r)]:
    lb(ws,r,1,lab)
    for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"=IF({cl}{rev_pnl}=0,0,{cl}{nr}/{cl}{rev_pnl})",FP)
    r+=1

# ===== CASH FLOW =====
ws=wb.create_sheet("Cash Flow");ws.sheet_properties.tabColor="00B0F0"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"CASH FLOW (M VND)");r+=2;time_hdrs(ws,r);r+=1

def pnl(row,i):return f"'P&L'!{get_column_letter(TSC+i)}{row}"

sc(ws,r,1,"Operating CF");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
ni_cf=r;lb(ws,r,1,"  Net income")
for i in range(NTC):fm(ws,r,TSC+i,f"={pnl(ni_r,i)}",FMT)
r+=1
da_cf=r;lb(ws,r,1,"  Add: D&A")
for i in range(NTC):fm(ws,r,TSC+i,f"=-{pnl(dep_pnl,i)}",FMT)
r+=1
wc_cf=r;lb(ws,r,1,"  Working capital chg")
for i in range(NTC):
    cl=get_column_letter(TSC+i);rv_f=pnl(rev_pnl,i)
    if i==0:fm(ws,r,TSC+i,f"=-{rv_f}*ar_days/365",FMT)
    else:fm(ws,r,TSC+i,f"=-({rv_f}-{pnl(rev_pnl,i-1)})*ar_days/365",FMT)
r+=1
ocf=r;lb(ws,r,1,"OPERATING CF")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ni_cf}+{cl}{da_cf}+{cl}{wc_cf}",FMT)
r+=2

sc(ws,r,1,"Investing CF");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
capex_cf=r;lb(ws,r,1,"  CAPEX")
for i in range(NTC):fm(ws,r,TSC+i,f"=-{cs(cx_tot_r,i)}",FMT)
r+=1
icf=r;lb(ws,r,1,"INVESTING CF")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{capex_cf}",FMT)
r+=2

fcf_r=r;lb(ws,r,1,"FREE CASH FLOW")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ocf}+{cl}{icf}",FMT)
r+=1
cum_r=r;lb(ws,r,1,"Cumulative CF")
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    if i==0:fm(ws,r,TSC+i,f"={cl}{fcf_r}",FMT)
    else:fm(ws,r,TSC+i,f"={get_column_letter(TSC+i-1)}{cum_r}+{cl}{fcf_r}",FMT)

# ===== DCF =====
ws=wb.create_sheet("DCF & Returns");ws.sheet_properties.tabColor="C00000"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=18
r=1;sc(ws,r,1,"DCF & RETURNS");r+=2
sc(ws,r,1,"Annual FCF");r+=1
hd(ws,r,1,"Year");hd(ws,r,2,"FCF (M VND)");r+=1
fs=r
for s,e,lbl in AN:
    lb(ws,r,1,lbl)
    fm(ws,r,2,f"=SUM('Cash Flow'!{get_column_letter(TSC+s)}{fcf_r}:{get_column_letter(TSC+e)}{fcf_r})",FMT)
    r+=1
fe=r-1;r+=1
sc(ws,r,1,"Metrics");r+=1
npv_r=r;lb(ws,r,1,"NPV");fm(ws,r,2,f"=NPV(discount_rate,B{fs+1}:B{fe})+B{fs}",FMT);r+=1
irr_r=r;lb(ws,r,1,"IRR");fm(ws,r,2,f'=IFERROR(IRR(B{fs}:B{fe}),"N/A")',FP);r+=1
inv_r2=r;lb(ws,r,1,"Max cash burn");fm(ws,r,2,f"=-MIN('Cash Flow'!{get_column_letter(TSC)}{cum_r}:'Cash Flow'!{get_column_letter(TSC+NTC-1)}{cum_r})",FMT);r+=1
lb(ws,r,1,"ROI");fm(ws,r,2,f"=IF(B{inv_r2}=0,0,('Cash Flow'!{get_column_letter(TSC+NTC-1)}{cum_r}+B{inv_r2})/B{inv_r2})",FP);r+=1
lb(ws,r,1,"CIT rate");fm(ws,r,2,"=cit_rate",FP);r+=1
lb(ws,r,1,"WACC");fm(ws,r,2,"=discount_rate",FP)

# Save
state['rev_tot_r']=rev_tot_r;state['cogs_tot_r']=cogs_tot_r
state['cx_tot_r']=cx_tot_r;state['ox_tot_r']=ox_tot_r;state['tot_cost_r']=tot_cost_r
state['rev_pnl']=rev_pnl;state['gp_r']=gp_r;state['ebitda_r']=ebitda_r;state['ni_r']=ni_r
state['dep_pnl']=dep_pnl;state['tax_r']=tax_r;state['opex_pnl']=opex_pnl
state['fcf_r']=fcf_r;state['cum_r']=cum_r;state['npv_r']=npv_r;state['irr_r']=irr_r
with open('/tmp/v4_state.json','w') as f:json.dump(state,f)
wb.save(OUT)
print("Part 3 done: Revenue + Cost + P&L + Cash Flow + DCF")
print(f"Sheets: {wb.sheetnames}")
