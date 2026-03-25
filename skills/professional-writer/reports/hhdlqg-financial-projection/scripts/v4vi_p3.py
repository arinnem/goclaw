"""
v4 Vietnamese — Part 3: Scenarios + Budget 2025/2026 + Guide + Glossary + finalize
"""
import openpyxl, json, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/v4vi_state.json') as f: state=json.load(f)
OUT=state['out'];FINAL=state['final']

HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LBF=Font(size=11);TF=Font(bold=True,size=11)
TFL=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
PH_FILL=PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid")
PH_FONT=Font(color="0000FF",size=11)
BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FMT='#,##0';FP='0.0%'
MN=['','Th1','Th2','Th3','Th4','Th5','Th6','Th7','Th8','Th9','Th10','Th11','Th12']
TSC=3;NTC=42

def hd(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def sc(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
def lb(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LBF;cl.border=BD
def fm(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=Font(size=11);cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def tot(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=TF;cl.fill=TFL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def text(ws,r,v,bold=False,color="333333"):
    c=ws.cell(row=r,column=1,value=v);c.font=Font(size=10,color=color,bold=bold);ws.merge_cells(f"A{r}:F{r}")

wb=openpyxl.load_workbook(OUT)

# ===== KỊCH BẢN =====
ws=wb.create_sheet("Kịch bản");ws.sheet_properties.tabColor="FFC000"
ws.column_dimensions['A'].width=32
for c in ['B','C','D','E','F']:ws.column_dimensions[c].width=16
r=1;sc(ws,r,1,"PHÂN TÍCH KỊCH BẢN");ws.merge_cells("A1:F1");r+=1
ws.cell(row=r,column=1,value="Sao chép giá trị từ cột kịch bản vào trang Assumptions để chạy.").font=Font(italic=True,size=10,color="666666");r+=2
for ci,h in[(1,"Tham số"),(2,"XẤU NHẤT"),(3,"THẬN TRỌNG"),(4,"CƠ SỞ"),(5,"LẠC QUAN"),(6,"TỐT NHẤT")]:hd(ws,r,ci,h)
r+=1
scen=[
    ("Trần DN (adoption_max)",200,600,"=adoption_max",2000,3500),
    ("Hệ số dốc S-curve",0.5,0.7,"=adoption_steepness",1.2,1.5),
    ("Tỷ lệ rời bỏ",0.35,0.25,"=churn_rate",0.15,0.10),
    ("Thị phần GMV",0.001,0.003,"=platform_market_share",0.01,0.02),
    ("Phí chứng nhận (Tr)",15,20,"=cert_fee_yr1",30,40),
    ("Tăng phí/năm",0.0,0.03,"=cert_fee_escalation",0.08,0.12),
    ("Phí giao dịch",0.01,0.015,"=tx_take_rate",0.025,0.03),
    ("Marketing cơ sở (Tr)",8000,7500,"=opex_marketing",6000,5000),
    ("WACC",0.20,0.18,"=discount_rate",0.12,0.10),
    ("Ngưỡng GĐ2",500,400,"=stage2_gate",200,150),
]
for label,w,be,ba,bu,best in scen:
    lb(ws,r,1,label)
    ip=isinstance(w,float) and w<1;f=FP if ip else FMT
    for ci,v,clr in[(2,w,"CC0000"),(3,be,None),(4,ba,None),(5,bu,None),(6,best,"00B050")]:
        c=ws.cell(row=r,column=ci,value=v);c.border=BD;c.number_format=f
        if clr:c.font=Font(size=11,color=clr)
        if ci==4:c.fill=TFL;c.font=Font(bold=True,size=11)
    r+=1

r+=2;sc(ws,r,1,"🔑 YẾU TỐ ẢNH HƯỞNG LỚN NHẤT ĐẾN NPV/IRR");ws.merge_cells(f"A{r}:F{r}");r+=2
text(ws,r,"3 yếu tố ảnh hưởng mạnh nhất đến kết quả mô hình:",bold=True);r+=2

text(ws,r,"1. TRẦN DOANH NGHIỆP (adoption_max) — YẾU TỐ SỐ 1",bold=True,color="1F4E79");r+=1
text(ws,r,"  Mọi nguồn doanh thu đều phụ thuộc số DN. Cert, đào tạo, kiểm toán — tỷ lệ thuận với DN.");r+=1
text(ws,r,"  Giai đoạn 2 (phí giao dịch, DaaS, QC) cũng phụ thuộc vượt ngưỡng. Tăng gấp đôi DN ≈ tăng gấp đôi DT.");r+=1
text(ws,r,"  → Xấu nhất: 200 = dự án không có sức hút. Tốt nhất: 3,500 = trở thành chuẩn ngành.");r+=2

text(ws,r,"2. THỊ PHẦN GMV (platform_market_share) — YẾU TỐ SỐ 2",bold=True,color="1F4E79");r+=1
text(ws,r,"  % GMV TMĐT du lịch VN chạy qua nền tảng VV. Thị trường ~82.500 tỷ, tăng 8%/năm.");r+=1
text(ws,r,"  Chênh 0.1% = hàng trăm tỷ VND chênh lệch GMV → hàng tỷ phí giao dịch.");r+=1
text(ws,r,"  → Xấu nhất: 0.1% = ngách. Tốt nhất: 2% = nền tảng thống lĩnh.");r+=2

text(ws,r,"3. TỶ LỆ RỜI BỎ (churn_rate) — YẾU TỐ SỐ 3",bold=True,color="1F4E79");r+=1
text(ws,r,"  % DN không gia hạn chứng nhận mỗi năm. Churn cao = phải liên tục thay thế DN đã mất.");r+=1
text(ws,r,"  35% churn: cần 35 DN mới/năm trên 100 DN chỉ để giữ nguyên. Rất tốn kém.");r+=1
text(ws,r,"  → Xấu nhất: 35% = chứng nhận không có giá trị. Tốt nhất: 10% = sản phẩm dính.");r+=2

sc(ws,r,1,"📖 MÔ TẢ CHI TIẾT TỪNG KỊCH BẢN");ws.merge_cells(f"A{r}:F{r}");r+=2

text(ws,r,"XẤU NHẤT — 'Ra mắt thất bại'",bold=True,color="CC0000");r+=1
text(ws,r,"  Chỉ 200 DN trong 5 năm. Thị trường không thấy giá trị chứng nhận.");r+=1
text(ws,r,"  GĐ2 không kích hoạt (200 < 300). Doanh thu chỉ từ phí cert, quy mô nhỏ.");r+=1
text(ws,r,"  Kết quả: Dự án bị đóng sau 2-3 năm. Tổng lỗ = CAPEX + chi phí vận hành.");r+=2

text(ws,r,"THẬN TRỌNG — 'Tăng trưởng chậm'",bold=True,color="996600");r+=1
text(ws,r,"  600 DN trong 5 năm. Chứng nhận được một số công nhận nhưng tăng chậm.");r+=1
text(ws,r,"  GĐ2 kích hoạt muộn, thị phần nhỏ 0.3%. Có thể hòa vốn 2029-2030.");r+=1
text(ws,r,"  Kết quả: Tồn tại nhưng cần tiếp tục bù lỗ. Lợi nhuận biên.");r+=2

text(ws,r,"CƠ SỞ — 'Tăng trưởng thực tế' (Giả định hiện tại)",bold=True,color="1F4E79");r+=1
text(ws,r,"  1.200 DN đến 2030 (3% SAM). Chứng nhận được thị trường công nhận.");r+=1
text(ws,r,"  GĐ2 kích hoạt 2027-2028, thị phần 0.5%. Hòa vốn khoảng 2027-2028.");r+=1
text(ws,r,"  Kết quả: Doanh nghiệp khả thi, tạo lợi nhuận vừa phải.");r+=2

text(ws,r,"LẠC QUAN — 'Tăng trưởng mạnh'",bold=True,color="0070C0");r+=1
text(ws,r,"  2.000 DN = 5% SAM. Chứng nhận VV trở thành yếu tố cạnh tranh.");r+=1
text(ws,r,"  GĐ2 kích hoạt sớm, thị phần 1%. Lợi nhuận tốt, thu hút đầu tư tiếp.");r+=1
text(ws,r,"  Kết quả: Doanh nghiệp hấp dẫn. Có thể mở rộng khu vực.");r+=2

text(ws,r,"TỐT NHẤT — 'Dẫn đầu thị trường'",bold=True,color="00B050");r+=1
text(ws,r,"  3.500 DN = chuẩn chứng nhận quốc gia cho du lịch Việt Nam.");r+=1
text(ws,r,"  Thị phần 2%, churn 10%. Chính phủ bắt buộc hoặc khuyến khích chứng nhận.");r+=1
text(ws,r,"  Kết quả: Sẵn sàng IPO hoặc mục tiêu M&A. Mở rộng Đông Nam Á.");r+=2

# ===== NGÂN SÁCH =====
def build_budget_vi(wb,year,months,title):
    ws=wb.create_sheet(title);ws.sheet_properties.tabColor="4472C4"
    ws.column_dimensions['A'].width=35;ws.column_dimensions['B'].width=12
    nm=len(months);tc=3+nm
    r=1;sc(ws,r,1,f"NGÂN SÁCH {year} (Tr VND)");r+=2
    hd(ws,r,1,"Hạng mục");hd(ws,r,2,"ĐVT")
    for j,m in enumerate(months):
        col=3+j;hd(ws,r,col,MN[m]);ws.column_dimensions[get_column_letter(col)].width=12
    hd(ws,r,tc,"TỔNG");ws.column_dimensions[get_column_letter(tc)].width=14;r+=1
    def rt(r_):return f"=SUM({get_column_letter(3)}{r_}:{get_column_letter(3+nm-1)}{r_})"
    ye=year-2025

    sc(ws,r,1,"DOANH THU");
    for j in range(nm):ws.cell(row=r,column=3+j).fill=SC_FILL
    ws.cell(row=r,column=tc).fill=SC_FILL;r+=1
    cert_r=r;lb(ws,r,1,"  Phí chứng nhận");lb(ws,r,2,"Tr VND")
    for j,m in enumerate(months):
        yf=year+(m-0.5)/12
        fm(ws,r,3+j,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({yf:.3f}-adoption_midpoint))))*cert_fee_yr1*(1+cert_fee_escalation)^{ye}/12",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    tr_r=r;lb(ws,r,1,"  Đào tạo");lb(ws,r,2,"Tr VND")
    for j,m in enumerate(months):
        yf=year+(m-0.5)/12
        fm(ws,r,3+j,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({yf:.3f}-adoption_midpoint))))*training_pct*training_fee/12",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    au_r=r;lb(ws,r,1,"  Kiểm toán");lb(ws,r,2,"Tr VND")
    for j,m in enumerate(months):
        yf=year+(m-0.5)/12
        fm(ws,r,3+j,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({yf:.3f}-adoption_midpoint))))*audit_fee/12",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    rev_r=r;lb(ws,r,1,"TỔNG DOANH THU")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"=SUM({cl}{cert_r}:{cl}{au_r})",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=2

    sc(ws,r,1,"CAPEX");
    for j in range(nm):ws.cell(row=r,column=3+j).fill=SC_FILL
    ws.cell(row=r,column=tc).fill=SC_FILL;r+=1
    cx_start=r
    if year==2025:
        for nm2,lbl in[("capex_cloud_setup","Cloud"),("capex_backend","Backend"),("capex_frontend","Frontend"),("capex_data_platform","Data"),("capex_security_setup","Bảo mật"),("capex_ux_design","UX/UI"),("capex_integration","Tích hợp"),("capex_devops","DevOps"),("capex_testing","QA"),("capex_contingency","Dự phòng")]:
            lb(ws,r,1,f"  {lbl}");lb(ws,r,2,"Tr VND")
            for j in range(nm):fm(ws,r,3+j,f"={nm2}/{nm}",FMT)
            tot(ws,r,tc,rt(r),FMT);r+=1
    else:
        lb(ws,r,1,"  Bảo trì");lb(ws,r,2,"Tr VND")
        for j in range(nm):fm(ws,r,3+j,"=capex_total*capex_maintenance_pct/12",FMT)
        tot(ws,r,tc,rt(r),FMT);r+=1
    cx_end=r-1;cx_r=r;lb(ws,r,1,"TỔNG CAPEX")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"=SUM({cl}{cx_start}:{cl}{cx_end})",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=2

    sc(ws,r,1,"OPEX");
    for j in range(nm):ws.cell(row=r,column=3+j).fill=SC_FILL
    ws.cell(row=r,column=tc).fill=SC_FILL;r+=1
    ox_start=r
    hc={2025:{"QL(2)":"salary_mgmt*2","PT(12)":"salary_dev*12","VHKD(2)":"salary_bizops*2"},
        2026:{"QL(2)":"salary_mgmt*2","PT(12)":"salary_dev*12","VHKD(6)":"salary_bizops*6","DL(2)":"salary_data*2"}}
    for dept,formula in hc.get(year,hc[2026]).items():
        lb(ws,r,1,f"  NS — {dept}");lb(ws,r,2,"Tr VND")
        for j in range(nm):fm(ws,r,3+j,f"={formula}*(1+salary_escalation)^{ye}*(1+bhxh_rate)",FMT)
        tot(ws,r,tc,rt(r),FMT);r+=1
    lb(ws,r,1,"  Tuyển dụng");lb(ws,r,2,"Tr VND")
    new_h=16 if year==2025 else 6
    for j in range(nm):
        if j==0:fm(ws,r,3+j,f"={new_h}*30*12*recruit_cost_pct",FMT)
        else:fm(ws,r,3+j,"=0",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    for lbl,formula in[("Marketing",f"=opex_marketing*(1+marketing_growth)^{ye}/12"),("Thương hiệu","=opex_branding/12"),
        ("Cloud","=opex_cloud_base/12"),("Bảo mật","=opex_security/12"),("Đối ngoại","=opex_external/12"),
        ("Pháp lý","=opex_legal/12"),("Đối tác","=opex_partner_mgmt/12")]:
        lb(ws,r,1,f"  {lbl}");lb(ws,r,2,"Tr VND")
        for j in range(nm):fm(ws,r,3+j,formula,FMT)
        tot(ws,r,tc,rt(r),FMT);r+=1
    if year==2025:
        lb(ws,r,1,"  Thí điểm");lb(ws,r,2,"Tr VND")
        for j in range(nm):fm(ws,r,3+j,"=opex_pilot_eval/6",FMT)
        tot(ws,r,tc,rt(r),FMT);r+=1
    lb(ws,r,1,"  Hỗ trợ KH");lb(ws,r,2,"Tr VND")
    for j,m in enumerate(months):
        yf=year+(m-0.5)/12
        fm(ws,r,3+j,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({yf:.3f}-adoption_midpoint))))*opex_support_per_ent/12",FMT)
    tot(ws,r,tc,rt(r),FMT);ox_end=r;r+=1
    ox_sub=r;lb(ws,r,1,"OPEX cộng dồn")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"=SUM({cl}{ox_start}:{cl}{ox_end})",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    cg=r;lb(ws,r,1,"  Dự phòng");lb(ws,r,2,"Tr VND")
    for j in range(nm):cl=get_column_letter(3+j);fm(ws,r,3+j,f"={cl}{ox_sub}*opex_contingency_pct",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    ox_tot=r;lb(ws,r,1,"TỔNG OPEX")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"={cl}{ox_sub}+{cl}{cg}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=2
    sc(ws,r,1,"TỔNG KẾT");
    for j in range(nm):ws.cell(row=r,column=3+j).fill=SC_FILL
    ws.cell(row=r,column=tc).fill=SC_FILL;r+=1
    sr=r;lb(ws,r,1,"Doanh thu")
    for j in range(nm):cl=get_column_letter(3+j);fm(ws,r,3+j,f"={cl}{rev_r}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    scx=r;lb(ws,r,1,"CAPEX")
    for j in range(nm):cl=get_column_letter(3+j);fm(ws,r,3+j,f"={cl}{cx_r}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    sox=r;lb(ws,r,1,"OPEX")
    for j in range(nm):cl=get_column_letter(3+j);fm(ws,r,3+j,f"={cl}{ox_tot}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    tb=r;lb(ws,r,1,"TỔNG NGÂN SÁCH")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"={cl}{scx}+{cl}{sox}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    nr=r;lb(ws,r,1,"RÒNG (DT - Ngân sách)")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"={cl}{sr}-{cl}{tb}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    lb(ws,r,1,"Lũy kế")
    for j in range(nm):
        cl=get_column_letter(3+j)
        if j==0:fm(ws,r,3+j,f"={cl}{nr}",FMT)
        else:fm(ws,r,3+j,f"={get_column_letter(3+j-1)}{r}+{cl}{nr}",FMT)
    tot(ws,r,tc,f"={get_column_letter(3+nm-1)}{r}",FMT)

build_budget_vi(wb,2025,list(range(7,13)),"Ngân sách 2025")
build_budget_vi(wb,2026,list(range(1,13)),"Ngân sách 2026")

# ===== HƯỚNG DẪN =====
ws=wb.create_sheet("Hướng dẫn",0)
ws.sheet_properties.tabColor="00B050"
ws.column_dimensions['A'].width=25;ws.column_dimensions['B'].width=55;ws.column_dimensions['C'].width=35
r=1;ws.cell(row=r,column=1,value="VISIT VIETNAM — MÔ HÌNH TÀI CHÍNH v4 — HƯỚNG DẪN").font=Font(bold=True,size=16,color="1F4E79")
ws.merge_cells("A1:C1");r+=1
ws.cell(row=r,column=1,value="v4.0 | 21/03/2026 | Phiên bản tiếng Việt").font=Font(italic=True,size=10,color="666666");r+=2

sc(ws,r,1,"DANH SÁCH CÁC TRANG");r+=1
hd(ws,r,1,"Trang");hd(ws,r,2,"Mục đích");hd(ws,r,3,"Ghi chú");r+=1
sheets=[
    ("Hướng dẫn","Giải thích mô hình, thuật ngữ, công thức, cách sử dụng","Đọc trước"),
    ("Assumptions","49 named ranges. MỌI đầu vào ở đây (ô vàng).","Chỉ sửa ô vàng"),
    ("Đầu tư ban đầu","10 hạng mục CAPEX → capex_total. Lịch khấu hao.","capex_total dùng khắp nơi"),
    ("Nhân sự","4 bộ phận × 42 kỳ. Lương = biên chế × named_salary × tăng × BHXH","Ô vàng = biên chế"),
    ("Quy mô thị trường","Đường cong S tăng trưởng DN. Year fracs từ base_year.","Không hardcode năm"),
    ("Doanh thu","6 nguồn thu. GĐ2 kích hoạt khi đạt ngưỡng stage2_gate.","Named ranges trong formulas"),
    ("Chi phí","CAPEX (Đầu tư) + OPEX (Nhân sự + 10 danh mục + dự phòng)","Liên kết chéo trang"),
    ("Kết quả KD","DT → Giá vốn → LN gộp → OPEX → EBITDA → KH → EBIT → Thuế 20% → LN ròng","Thuế TNDN 20%"),
    ("Dòng tiền","ĐTHĐ + ĐTĐT → DTTD + Lũy kế. Vốn lưu động qua ar_days.","Named ranges"),
    ("Thẩm định","FCF theo năm, NPV, IRR, ROI. Dùng discount_rate.","Công thức NPV duy nhất"),
    ("Kịch bản","5 kịch bản với giải thích chi tiết + yếu tố ảnh hưởng lớn nhất.","Copy vào Assumptions chạy"),
    ("Ngân sách 2025","Ngân sách tháng 7-12/2025. DT + CAPEX + OPEX + Dự phòng.","Sẵn sàng duyệt"),
    ("Ngân sách 2026","Ngân sách tháng 1-12/2026. Cùng cấu trúc, NS tăng.","Sẵn sàng duyệt"),
]
for n,p,nt in sheets:lb(ws,r,1,n);lb(ws,r,2,p);lb(ws,r,3,nt);r+=1
r+=1

# Glossary
sc(ws,r,1,"THUẬT NGỮ");r+=1
hd(ws,r,1,"Thuật ngữ");hd(ws,r,2,"Tên đầy đủ / Ý nghĩa");hd(ws,r,3,"Trong mô hình này");r+=1
terms=[
    ("NPV","Giá trị hiện tại ròng","Tổng dòng tiền tương lai chiết khấu về hiện tại. Dương = tạo giá trị."),
    ("IRR","Tỷ suất hoàn vốn nội bộ","Tỷ suất chiết khấu khi NPV = 0. Cao hơn = lợi nhuận tốt hơn."),
    ("WACC","Chi phí vốn bình quân gia quyền","Lợi nhuận tối thiểu NĐT kỳ vọng. Dùng làm discount_rate (15% cơ sở)."),
    ("ROI","Lợi nhuận trên đầu tư","(Dòng tiền lũy kế + Đầu tư) / Đầu tư."),
    ("EBITDA","LN trước lãi, thuế, khấu hao","Lợi nhuận hoạt động trước chi phí phi tiền mặt và thuế."),
    ("EBIT","LN trước lãi và thuế","= EBITDA - Khấu hao. Lợi nhuận trước thuế."),
    ("Thuế TNDN","Thuế thu nhập doanh nghiệp","Thuế suất chuẩn VN = 20%. Chỉ áp dụng khi EBIT > 0."),
    ("FCF","Dòng tiền tự do","= ĐTHĐ + ĐTĐT. Tiền thực sự có sẵn sau mọi chi tiêu."),
    ("CAPEX","Chi phí đầu tư","Đầu tư 1 lần vào tài sản (nền tảng 5.280Tr). Khấu hao dần."),
    ("OPEX","Chi phí hoạt động","Chi phí liên tục: NS, marketing, cloud, pháp lý... Ghi nhận ngay."),
    ("COGS","Giá vốn hàng bán","Chi phí trực tiếp gắn với DT: phí thanh toán (2.5% DT giao dịch)."),
    ("BHXH","Bảo hiểm xã hội","DN đóng ~21.5%: BHXH 17.5% + BHYT 3% + BHTN 1%."),
    ("TAM","Tổng thị trường tiềm năng","Tất cả DN du lịch VN (~40.000)."),
    ("SAM","Thị trường phục vụ được","DN hoạt động số, có thể chứng nhận (~15.000)."),
    ("GMV","Tổng giá trị giao dịch","Tổng giá trị giao dịch qua nền tảng."),
    ("VCR","Bản ghi du khách chứng nhận","Dữ liệu tạo ra/DN. Cơ sở tính giá DaaS."),
    ("S-curve","Đường cong tăng trưởng logistic","DN tăng theo mẫu chậm→nhanh→chậm. max/(1+EXP(-k*(t-midpoint)))"),
    ("Giai đoạn 1","Nền tảng chứng nhận (2025-2027)","DT từ phí cert, đào tạo, kiểm toán. Không cần khối lượng GD."),
    ("Giai đoạn 2","Nền tảng giao dịch (2028-2030)","DT từ phí GD, DaaS, QC. Yêu cầu ≥300 DN (stage2_gate)."),
    ("Named Range","Tên được đặt cho ô Excel","VD: cert_fee_yr1 → Assumptions!$B$26. Dùng trong công thức."),
    ("Churn","Tỷ lệ rời bỏ","% DN không gia hạn mỗi năm. Giảm net_enterprises."),
]
for t,full,ctx in terms:lb(ws,r,1,t);lb(ws,r,2,full);lb(ws,r,3,ctx);r+=1

# Reorder
order=["Hướng dẫn","Assumptions","Đầu tư ban đầu","Nhân sự","Quy mô thị trường","Doanh thu","Chi phí","Kết quả KD","Dòng tiền","Thẩm định","Kịch bản","Ngân sách 2025","Ngân sách 2026"]
for i,n in enumerate(order):
    if n in wb.sheetnames:wb.move_sheet(n,offset=i-wb.sheetnames.index(n))

wb.save(OUT)
print(f"v4 VN complete: {OUT}")
print(f"Sheets: {wb.sheetnames}")

try:shutil.copy2(OUT,FINAL);print(f"Copied to: {FINAL}")
except Exception as e:print(f"Copy failed: {e}\nFile at: {OUT}")
