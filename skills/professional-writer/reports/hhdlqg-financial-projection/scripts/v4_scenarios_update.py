"""
Add detailed scenario explanations to the Scenarios sheet in v4.
Load existing file, modify only Scenarios sheet, save.
"""
import openpyxl, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\hhdlqg-financial-projection\VisitVietnam_FinancialModel_v4.xlsx"
TMP = r"E:\tmp\v4_scenarios_update.xlsx"
shutil.copy2(SRC, TMP)
wb = openpyxl.load_workbook(TMP)

ws = wb['Scenarios']
BD = Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),
    top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
SC_FILL = PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT = Font(bold=True,size=11,color="1F4E79")
HD_FILL = PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT = Font(color="FFFFFF",bold=True,size=11)

# Find the last used row
r = ws.max_row + 3

def section(title):
    global r
    c = ws.cell(row=r, column=1, value=title); c.font=SC_FONT; c.fill=SC_FILL
    ws.merge_cells(f"A{r}:F{r}")
    for col in range(2,7): ws.cell(row=r,column=col).fill=SC_FILL
    r += 1

def text(content, bold=False, color="333333", indent=0):
    global r
    prefix = "  " * indent
    c = ws.cell(row=r, column=1, value=prefix + content)
    c.font = Font(size=10, color=color, bold=bold)
    ws.merge_cells(f"A{r}:F{r}")
    r += 1

# ===== WHAT DRIVES RESULTS THE MOST =====
section("🔑 KEY DRIVERS — WHAT AFFECTS NPV/IRR THE MOST")
r += 1

text("The model's NPV and IRR are most sensitive to these 3 factors, in order of impact:", bold=True)
r += 1

text("1. ADOPTION MAX (adoption_max) — #1 DRIVER", bold=True, color="1F4E79")
text("This is the S-curve ceiling: maximum enterprises that will ever certify.", indent=1)
text("Every revenue stream depends on enterprise count. Cert fees, training, audit — all linear to enterprises.", indent=1)
text("Stage 2 (transaction fees, DaaS, ads) also depends on crossing the gate threshold.", indent=1)
text("Doubling adoption_max roughly doubles total revenue while costs only grow ~30-40%.", indent=1)
text("→ Worst: 200 = project never gains traction. Best: 3,500 = becomes the standard.", indent=1)
r += 1

text("2. PLATFORM MARKET SHARE (platform_market_share) — #2 DRIVER", bold=True, color="1F4E79")
text("% of Vietnam's e-tourism GMV that flows through the VV platform.", indent=1)
text("Vietnam e-tourism = ~82,500 Bn VND (2025), growing 8%/yr → ~121,000 Bn by 2030.", indent=1)
text("Even 0.1% difference = 121 Bn VND difference in GMV → 2.4 Bn in tx fees at 2% take rate.", indent=1)
text("This only activates in Stage 2 (after hitting gate), so impact shows in later years.", indent=1)
text("→ Worst: 0.1% = niche player. Best: 2% = dominant platform (similar to v2's broken assumption).", indent=1)
r += 1

text("3. CHURN RATE (churn_rate) — #3 DRIVER", bold=True, color="1F4E79")
text("% of certified enterprises that don't renew each year.", indent=1)
text("High churn means constantly replacing lost enterprises — growth slows, costs don't drop.", indent=1)
text("At 35% churn: need 35 new signups/yr per 100 just to stay flat. Very expensive.", indent=1)
text("At 10% churn: compound growth accelerates, LTV per enterprise increases.", indent=1)
text("→ Worst: 35% = certification has low perceived value. Best: 10% = sticky, high-value.", indent=1)
r += 2

# ===== SCENARIO NARRATIVES =====
section("📖 DETAILED SCENARIO NARRATIVES")
r += 1

text("WORST CASE — 'Failed Launch'", bold=True, color="CC0000")
text("Only 200 enterprises adopt over 5 years. Market sees no value in certification.", indent=1)
text("Stage 2 never activates (200 < 300 gate). Revenue = cert fees only, from tiny base.", indent=1)
text("High churn 35% means even the few adopters leave. High marketing spend with no return.", indent=1)
text("Discount rate 20% reflects extreme risk. Project is a net loss.", indent=1)
text("Likely outcome: Project shut down after 2-3 years. Total loss = initial CAPEX + operating losses.", indent=1)
r += 1

text("BEAR CASE — 'Slow Burn'", bold=True, color="996600")
text("600 enterprises over 5 years. Certification gains some recognition but growth is slow.", indent=1)
text("Stage 2 activates late (600 > 300 gate) but with only 0.3% market share, modest tx revenue.", indent=1)
text("25% churn means 1 in 4 enterprises don't renew — need strong sales to maintain.", indent=1)
text("Model may break even operationally by 2029-2030 but NPV is near zero or slightly negative.", indent=1)
text("Likely outcome: Survives but requires ongoing subsidy. Marginal returns.", indent=1)
r += 1

text("BASE CASE — 'Realistic Growth' (Current Assumptions)", bold=True, color="1F4E79")
text("1,200 enterprises by 2030 (3% of SAM). Certification becomes recognized in the market.", indent=1)
text("Stage 2 activates in 2027-2028 with 0.5% market share — meaningful but not dominant.", indent=1)
text("20% churn is realistic for B2B subscription services in Vietnam.", indent=1)
text("WACC 15% reflects startup risk in a developing market.", indent=1)
text("Breakeven expected around 2027-2028. Positive cumulative CF by 2029-2030.", indent=1)
text("Likely outcome: Viable business generating moderate returns.", indent=1)
r += 1

text("BULL CASE — 'Strong Adoption'", bold=True, color="0070C0")
text("2,000 enterprises = 5% of SAM. VV certification becomes a market differentiator.", indent=1)
text("Stage 2 activates early with 1% market share. Transaction revenue becomes significant.", indent=1)
text("15% churn shows certification has real value. Enterprises renew because it drives business.", indent=1)
text("Lower WACC 12% reflects proven business model.", indent=1)
text("Strong returns with meaningful IRR. Could attract follow-on investment.", indent=1)
text("Likely outcome: Attractive business. Consider geographic expansion.", indent=1)
r += 1

text("BEST CASE — 'Market Leader'", bold=True, color="00B050")
text("3,500 enterprises = 8.75% of SAM. VV becomes THE certification standard for Vietnam tourism.", indent=1)
text("2% market share in e-tourism = dominant platform position.", indent=1)
text("10% churn = extremely sticky product. Enterprises can't afford NOT to have it.", indent=1)
text("Government adoption or mandated certification would push toward this scenario.", indent=1)
text("Outstanding returns. Platform effects create a moat.", indent=1)
text("Likely outcome: IPO-ready or acquisition target. Regional expansion to SE Asia.", indent=1)
r += 2

# ===== SECONDARY DRIVERS =====
section("📊 SECONDARY DRIVERS (moderate impact)")
r += 1

text("4. Cert fee level (cert_fee_yr1) — direct impact on Stage 1 revenue")
text("   ↳ 15M vs 40M = 2.67x revenue difference, but too high → increases churn")
text("5. Cert fee escalation — compound effect over 5 years")
text("   ↳ 0% vs 12% = at 1000 ent, difference of 2.3B/yr by 2030")
text("6. Marketing spend — indirect: drives adoption speed, but is a cost")
text("   ↳ 5000M vs 8000M/yr = 3B/yr cost difference. Must justify with faster adoption")
text("7. Discount rate — changes present value of future cash flows")
text("   ↳ 10% vs 20% = can flip NPV from positive to negative on same cash flows")
text("8. Stage 2 gate — when transaction revenue kicks in")
text("   ↳ 150 vs 500 gate = 1-2 year difference in Stage 2 activation")

wb.save(SRC)
print(f"Scenarios sheet updated in: {SRC}")
