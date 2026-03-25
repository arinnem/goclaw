"""
v4 Vietnamese — Part 2: Market Sizing + Revenue + Cost + P&L + Cash Flow + DCF
"""
import openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/v4vi_state.json') as f: state=json.load(f)
OUT=state['out'];hr_tot_r=state['hr_tot_r']

HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LBF=Font(size=11);TF=Font(bold=True,size=11)
TFL=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FMT='#,##0';FP='0.0%';FD='0.000'

def hd(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def sc(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
def lb(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LBF;cl.border=BD
def fm(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=Font(size=11);cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def tot(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=TF;cl.fill=TFL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)

MN=['','Th1','Th2','Th3','Th4','Th5','Th6','Th7','Th8','Th9','Th10','Th11','Th12']
TC=[]
for m in range(7,13):TC.append((f"{MN[m]} 2025",2025,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2026",2026,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2027",2027,'m',m))
for yr in [2028,2029,2030]:
    for q in range(1,5):TC.append((f"Q{q} {yr}",yr,'q',q))
TSC=3;NTC=42

def thdr(ws,r,c1="Hạng mục",c2="ĐVT"):
    hd(ws,r,1,c1);hd(ws,r,2,c2)
    for i,(l,y,t,m) in enumerate(TC):col=TSC+i;hd(ws,r,col,l);ws.column_dimensions[get_column_letter(col)].width=11

wb=openpyxl.load_workbook(OUT)

# ===== QUY MÔ THỊ TRƯỜNG =====
ws=wb.create_sheet("Quy mô thị trường");ws.sheet_properties.tabColor="4472C4"
ws.column_dimensions['A'].width=35;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"QUY MÔ THỊ TRƯỜNG & TĂNG TRƯỞNG");r+=2
idx_r=r;lb(ws,r,1,"Chỉ số kỳ");ws.cell(row=r,column=2,value="idx").font=LBF
for i in range(NTC):fm(ws,r,TSC+i,i,FMT)
r+=1
yfrac_r=r;lb(ws,r,1,"Năm phân số");ws.cell(row=r,column=2,value="năm").font=LBF
for i,(l,yr,t,m) in enumerate(TC):
    col=TSC+i;cl=get_column_letter(col)
    if t=='m':fm(ws,r,col,f"=base_year+(start_month-1+{cl}{idx_r})/12",FD)
    else:
        qm=(i-30)*3+1;fm(ws,r,col,f"=2028+({qm}-0.5)/12",FD)
r+=1
ybase_r=r;lb(ws,r,1,"Năm từ gốc");ws.cell(row=r,column=2,value="năm").font=LBF
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"=FLOOR({cl}{yfrac_r}-base_year,1)",FMT)
r+=1
div_r=r;lb(ws,r,1,"Ước số kỳ");ws.cell(row=r,column=2,value="chia").font=LBF
for i,(l,yr,t,m) in enumerate(TC):fm(ws,r,TSC+i,12 if t=='m' else 4,FMT)
r+=1
mon_r=r;lb(ws,r,1,"Số tháng/kỳ");ws.cell(row=r,column=2,value="tháng").font=LBF
for i,(l,yr,t,m) in enumerate(TC):fm(ws,r,TSC+i,1 if t=='m' else 3,FMT)
r+=2;thdr(ws,r);r+=1

ecom_r=r;lb(ws,r,1,"TT TMĐT du lịch");ws.cell(row=r,column=2,value="Tỷ VND").font=LBF
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"=ecom_market_base*(1+ecom_growth_rate)^{cl}{ybase_r}/{cl}{div_r}",'#,##0.0')
r+=2;sc(ws,r,1,"TĂNG TRƯỞNG DN (S-CURVE)")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1

def ms(row,i):return f"'Quy mô thị trường'!{get_column_letter(TSC+i)}{row}"

gross_r=r;lb(ws,r,1,"DN thô (S-curve)");ws.cell(row=r,column=2,value="DN").font=LBF
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({cl}{yfrac_r}-adoption_midpoint))))",FMT)
r+=1
churn_r=r;lb(ws,r,1,"DN rời bỏ");ws.cell(row=r,column=2,value="DN").font=LBF
for i in range(NTC):cl=get_column_letter(TSC+i)
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    if i==0:fm(ws,r,TSC+i,"=0",FMT)
    else:fm(ws,r,TSC+i,f"=ROUND({cl}{gross_r}*churn_rate/{cl}{div_r},0)",FMT)
r+=1
net_r=r;lb(ws,r,1,"DN hoạt động ròng");ws.cell(row=r,column=2,value="DN").font=LBF
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{gross_r}-{cl}{churn_r}",FMT)
r+=1
lb(ws,r,1,"DN mới/kỳ");ws.cell(row=r,column=2,value="DN").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    if i==0:fm(ws,r,TSC+i,f"={cl}{net_r}",FMT)
    else:fm(ws,r,TSC+i,f"=MAX(0,{cl}{net_r}-{get_column_letter(TSC+i-1)}{net_r})",FMT)
r+=1
vcr_r=r;lb(ws,r,1,"VCR (triệu bản ghi)");ws.cell(row=r,column=2,value="Tr").font=LBF
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"={cl}{net_r}*vcr_per_enterprise*{cl}{mon_r}/1000000",'0.00')

# ===== DOANH THU =====
ws=wb.create_sheet("Doanh thu");ws.sheet_properties.tabColor="00B050"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"BẢNG DOANH THU (Tr VND)");r+=2;thdr(ws,r);r+=1

sc(ws,r,1,"GIAI ĐOẠN 1")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
cert_r=r;lb(ws,r,1,"  Phí chứng nhận");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={ms(net_r,i)}*cert_fee_yr1*(1+cert_fee_escalation)^{ms(ybase_r,i)}/{ms(div_r,i)}",FMT)
r+=1
train_r=r;lb(ws,r,1,"  Đào tạo & tư vấn");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={ms(net_r,i)}*training_pct*training_fee/{ms(div_r,i)}",FMT)
r+=1
audit_r=r;lb(ws,r,1,"  Phí kiểm toán");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={ms(net_r,i)}*audit_fee/{ms(div_r,i)}",FMT)
r+=1
s1_r=r;lb(ws,r,1,"Tổng GĐ1")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"=SUM({cl}{cert_r}:{cl}{audit_r})",FMT)
r+=2;sc(ws,r,1,"GIAI ĐOẠN 2 (có điều kiện)")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
tx_r=r;lb(ws,r,1,"  Phí giao dịch");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=IF({ms(net_r,i)}>=stage2_gate,{ms(ecom_r,i)}*platform_market_share*tx_take_rate*1000,0)",FMT)
r+=1
daas_r=r;lb(ws,r,1,"  DaaS");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=IF({ms(net_r,i)}>=stage2_gate,{ms(vcr_r,i)}*daas_price,0)",FMT)
r+=1
ad_r=r;lb(ws,r,1,"  Quảng cáo");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=IF({ms(net_r,i)}>=stage2_gate,ad_impressions_base*{ms(mon_r,i)}*{ms(net_r,i)}/1000*ad_cpm/1000,0)",FMT)
r+=1
s2_r=r;lb(ws,r,1,"Tổng GĐ2")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"=SUM({cl}{tx_r}:{cl}{ad_r})",FMT)
r+=2
rev_tot_r=r;lb(ws,r,1,"TỔNG DOANH THU")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{s1_r}+{cl}{s2_r}",FMT)
r+=2;sc(ws,r,1,"GIÁ VỐN")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1;cogs_r=r;lb(ws,r,1,"  Phí thanh toán");lb(ws,r,2,"Tr VND")
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"={cl}{tx_r}*payment_proc_pct",FMT)
r+=1;cogs_tot_r=r;lb(ws,r,1,"TỔNG GIÁ VỐN")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{cogs_r}",FMT)

# ===== CHI PHÍ =====
ws=wb.create_sheet("Chi phí");ws.sheet_properties.tabColor="FF6600"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"BẢNG CHI PHÍ (Tr VND)");r+=2;thdr(ws,r);r+=1
sc(ws,r,1,"CAPEX");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
cdev_r=r;lb(ws,r,1,"  Phát triển nền tảng");lb(ws,r,2,"Tr VND")
for i,(l,yr,t,m) in enumerate(TC):
    if i<6:fm(ws,r,TSC+i,"=capex_total/6",FMT)
    else:fm(ws,r,TSC+i,"=0",FMT)
r+=1;cmnt_r=r;lb(ws,r,1,"  Bảo trì");lb(ws,r,2,"Tr VND")
for i,(l,yr,t,m) in enumerate(TC):
    if yr<=2025:fm(ws,r,TSC+i,"=0",FMT)
    else:fm(ws,r,TSC+i,f"=capex_total*capex_maintenance_pct/{ms(div_r,i)}",FMT)
r+=1;cx_tot_r=r;lb(ws,r,1,"TỔNG CAPEX")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{cdev_r}+{cl}{cmnt_r}",FMT)
r+=2;sc(ws,r,1,"OPEX");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
def hr_c(row,i):return f"'Nhân sự'!{get_column_letter(TSC+i)}{row}"
opex_start=r
lb(ws,r,1,"  Nhân sự (từ Nhân sự)");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={hr_c(hr_tot_r,i)}",FMT)
r+=1
opex_items = [
    ("  Marketing", lambda i: f"=opex_marketing*(1+marketing_growth)^{ms(ybase_r,i)}/{ms(div_r,i)}"),
    ("  Thương hiệu & sự kiện", lambda i: f"=opex_branding/{ms(div_r,i)}"),
    ("  Cloud & hạ tầng", lambda i: f"=opex_cloud_base/{ms(div_r,i)}+{ms(vcr_r,i)}*cloud_per_1M_vcr"),
    ("  Bảo mật", lambda i: f"=opex_security/{ms(div_r,i)}"),
    ("  Đối ngoại", lambda i: f"=opex_external/{ms(div_r,i)}"),
    ("  Pháp lý", lambda i: f"=opex_legal/{ms(div_r,i)}"),
]
for lbl, fn in opex_items:
    lb(ws,r,1,lbl);lb(ws,r,2,"Tr VND")
    for i in range(NTC):fm(ws,r,TSC+i,fn(i),FMT)
    r+=1
pilot_vi_r=r;lb(ws,r,1,"  Đánh giá thí điểm (Y1)");lb(ws,r,2,"Tr VND")
for i,(l,yr,t,m) in enumerate(TC):
    if yr==2025:fm(ws,r,TSC+i,f"=opex_pilot_eval/{ms(div_r,i)}",FMT)
    else:fm(ws,r,TSC+i,"=0",FMT)
r+=1;lb(ws,r,1,"  Quản lý đối tác");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"=opex_partner_mgmt/{ms(div_r,i)}",FMT)
r+=1;lb(ws,r,1,"  Hỗ trợ khách hàng");lb(ws,r,2,"Tr VND")
for i in range(NTC):fm(ws,r,TSC+i,f"={ms(net_r,i)}*opex_support_per_ent/{ms(div_r,i)}",FMT)
opex_end=r;r+=1
ox_sub_r=r;lb(ws,r,1,"OPEX cộng dồn")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"=SUM({cl}{opex_start}:{cl}{opex_end})",FMT)
r+=1;cont_r=r;lb(ws,r,1,"  Dự phòng");lb(ws,r,2,"Tr VND")
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"={cl}{ox_sub_r}*opex_contingency_pct",FMT)
r+=1;ox_tot_r=r;lb(ws,r,1,"TỔNG OPEX")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ox_sub_r}+{cl}{cont_r}",FMT)
r+=2;lb(ws,r,1,"TỔNG CHI PHÍ")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{cx_tot_r}+{cl}{ox_tot_r}",FMT)

# ===== KẾT QUẢ KINH DOANH =====
ws=wb.create_sheet("Kết quả KD");ws.sheet_properties.tabColor="7030A0"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"BÁO CÁO KẾT QUẢ KINH DOANH (Tr VND)");r+=2;thdr(ws,r);r+=1
def rv(row,i):return f"'Doanh thu'!{get_column_letter(TSC+i)}{row}"
def cs(row,i):return f"'Chi phí'!{get_column_letter(TSC+i)}{row}"

rev_pnl=r;lb(ws,r,1,"Doanh thu")
for i in range(NTC):fm(ws,r,TSC+i,f"={rv(rev_tot_r,i)}",FMT)
r+=1;cogs_pnl=r;lb(ws,r,1,"Giá vốn")
for i in range(NTC):fm(ws,r,TSC+i,f"=-{rv(cogs_tot_r,i)}",FMT)
r+=1;gp_r=r;lb(ws,r,1,"LỢI NHUẬN GỘP")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{rev_pnl}+{cl}{cogs_pnl}",FMT)
r+=2;opex_pnl=r;lb(ws,r,1,"Chi phí hoạt động")
for i in range(NTC):fm(ws,r,TSC+i,f"=-{cs(ox_tot_r,i)}",FMT)
r+=1;ebitda_r=r;lb(ws,r,1,"EBITDA")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{gp_r}+{cl}{opex_pnl}",FMT)
r+=1;dep_pnl=r;lb(ws,r,1,"Khấu hao")
for i,(l,yr,t,m) in enumerate(TC):
    if yr<=2029:fm(ws,r,TSC+i,f"=-capex_total/depreciation_years/{ms(div_r,i)}",FMT)
    else:fm(ws,r,TSC+i,"=0",FMT)
r+=1;ebit_r=r;lb(ws,r,1,"EBIT")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ebitda_r}+{cl}{dep_pnl}",FMT)
r+=1;tax_r=r;lb(ws,r,1,"Thuế TNDN (20%)")
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"=IF({cl}{ebit_r}>0,-{cl}{ebit_r}*cit_rate,0)",FMT)
r+=1;ni_r=r;lb(ws,r,1,"LỢI NHUẬN SAU THUẾ")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ebit_r}+{cl}{tax_r}",FMT)
r+=2;sc(ws,r,1,"BIÊN LỢI NHUẬN");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
for lab,nr in[("Biên gộp",gp_r),("Biên EBITDA",ebitda_r),("Biên ròng",ni_r)]:
    lb(ws,r,1,lab)
    for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"=IF({cl}{rev_pnl}=0,0,{cl}{nr}/{cl}{rev_pnl})",FP)
    r+=1

# ===== DÒNG TIỀN =====
ws=wb.create_sheet("Dòng tiền");ws.sheet_properties.tabColor="00B0F0"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"BÁO CÁO DÒNG TIỀN (Tr VND)");r+=2;thdr(ws,r);r+=1
def pnl(row,i):return f"'Kết quả KD'!{get_column_letter(TSC+i)}{row}"

sc(ws,r,1,"Dòng tiền hoạt động");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1;ni_cf=r;lb(ws,r,1,"  Lợi nhuận ròng")
for i in range(NTC):fm(ws,r,TSC+i,f"={pnl(ni_r,i)}",FMT)
r+=1;da_cf=r;lb(ws,r,1,"  Cộng: Khấu hao")
for i in range(NTC):fm(ws,r,TSC+i,f"=-{pnl(dep_pnl,i)}",FMT)
r+=1;wc_cf=r;lb(ws,r,1,"  Thay đổi vốn lưu động")
for i in range(NTC):
    rv_f=pnl(rev_pnl,i)
    if i==0:fm(ws,r,TSC+i,f"=-{rv_f}*ar_days/365",FMT)
    else:fm(ws,r,TSC+i,f"=-({rv_f}-{pnl(rev_pnl,i-1)})*ar_days/365",FMT)
r+=1;ocf=r;lb(ws,r,1,"DÒNG TIỀN HOẠT ĐỘNG")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ni_cf}+{cl}{da_cf}+{cl}{wc_cf}",FMT)
r+=2;sc(ws,r,1,"Dòng tiền đầu tư");
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1;capex_cf=r;lb(ws,r,1,"  Đầu tư CAPEX")
for i in range(NTC):fm(ws,r,TSC+i,f"=-{cs(cx_tot_r,i)}",FMT)
r+=1;icf=r;lb(ws,r,1,"DÒNG TIỀN ĐẦU TƯ")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{capex_cf}",FMT)
r+=2;fcf_r=r;lb(ws,r,1,"DÒNG TIỀN TỰ DO")
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{ocf}+{cl}{icf}",FMT)
r+=1;cum_r=r;lb(ws,r,1,"Dòng tiền lũy kế")
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    if i==0:fm(ws,r,TSC+i,f"={cl}{fcf_r}",FMT)
    else:fm(ws,r,TSC+i,f"={get_column_letter(TSC+i-1)}{cum_r}+{cl}{fcf_r}",FMT)

# ===== THẨM ĐỊNH =====
AN=[(0,5,'H2_2025'),(6,17,'2026'),(18,29,'2027'),(30,33,'2028'),(34,37,'2029'),(38,41,'2030')]
ws=wb.create_sheet("Thẩm định");ws.sheet_properties.tabColor="C00000"
ws.column_dimensions['A'].width=38;ws.column_dimensions['B'].width=18
r=1;sc(ws,r,1,"THẨM ĐỊNH DỰ ÁN & LỢI NHUẬN");r+=2
sc(ws,r,1,"Dòng tiền tự do theo năm");r+=1
hd(ws,r,1,"Năm");hd(ws,r,2,"FCF (Tr VND)");r+=1
fs=r
for s,e,lbl in AN:
    lb(ws,r,1,lbl);fm(ws,r,2,f"=SUM('Dòng tiền'!{get_column_letter(TSC+s)}{fcf_r}:{get_column_letter(TSC+e)}{fcf_r})",FMT);r+=1
fe=r-1;r+=1
sc(ws,r,1,"Chỉ số chính");r+=1
lb(ws,r,1,"NPV (Giá trị hiện tại ròng)");fm(ws,r,2,f"=NPV(discount_rate,B{fs+1}:B{fe})+B{fs}",FMT);r+=1
lb(ws,r,1,"IRR (Tỷ suất hoàn vốn nội bộ)");fm(ws,r,2,f'=IFERROR(IRR(B{fs}:B{fe}),"N/A")',FP);r+=1
inv2=r;lb(ws,r,1,"Đốt tiền tối đa");fm(ws,r,2,f"=-MIN('Dòng tiền'!{get_column_letter(TSC)}{cum_r}:'Dòng tiền'!{get_column_letter(TSC+NTC-1)}{cum_r})",FMT);r+=1
lb(ws,r,1,"ROI (Lợi nhuận đầu tư)");fm(ws,r,2,f"=IF(B{inv2}=0,0,('Dòng tiền'!{get_column_letter(TSC+NTC-1)}{cum_r}+B{inv2})/B{inv2})",FP);r+=1
lb(ws,r,1,"Thuế TNDN áp dụng");fm(ws,r,2,"=cit_rate",FP);r+=1
lb(ws,r,1,"Tỷ suất chiết khấu");fm(ws,r,2,"=discount_rate",FP)

# Save state for Part 3
state['rev_tot_r']=rev_tot_r;state['cogs_tot_r']=cogs_tot_r
state['cx_tot_r']=cx_tot_r;state['ox_tot_r']=ox_tot_r
state['rev_pnl']=rev_pnl;state['ni_r']=ni_r;state['dep_pnl']=dep_pnl
state['fcf_r']=fcf_r;state['cum_r']=cum_r
state['net_r']=net_r;state['vcr_r']=vcr_r;state['ecom_r']=ecom_r
state['div_r']=div_r;state['mon_r']=mon_r;state['ybase_r']=ybase_r;state['yfrac_r']=yfrac_r
with open('/tmp/v4vi_state.json','w') as f:json.dump(state,f)
wb.save(OUT)
print("Part 2 done: Market + Revenue + Cost + P&L + Cash Flow + DCF")
print(f"Sheets: {wb.sheetnames}")
