"""
v4 Vietnamese — Part 1: Assumptions + Initial Investment + HR Plan
Same structure/formulas as English v4, all text in Vietnamese.
"""
import openpyxl, json, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

PH_FILL=PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid")
PH_FONT=Font(color="0000FF",size=11)
HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LB=Font(size=11);TF=Font(bold=True,size=11)
TFL=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
NOTE=Font(italic=True,size=10,color="888888")
FMT_MN='#,##0';FMT_PCT='0.0%';FMT_INT='#,##0';FMT_DEC='0.0';FMT='#,##0';FD='0.000'

OUT=r"E:\tmp\v4_vi_build.xlsx"
FINAL=r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\hhdlqg-financial-projection\VisitVietnam_MoHinhTaiChinh_v4.xlsx"

wb=openpyxl.Workbook()
named={}
def add_nm(name,row):
    ref=f"Assumptions!$B${row}"
    dn=DefinedName(name,attr_text=ref);wb.defined_names.add(dn);named[name]=row

def ph(ws,r,c,v,fmt=None):cl=ws.cell(row=r,column=c,value=v);cl.font=PH_FONT;cl.fill=PH_FILL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def hd(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def sc(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
def lb(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LB;cl.border=BD
def fm(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=Font(size=11);cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def tot(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=TF;cl.fill=TFL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)

# ===== GIẢ ĐỊNH =====
ws=wb.active;ws.title="Assumptions";ws.sheet_properties.tabColor="FFD700"
ws.column_dimensions['A'].width=42;ws.column_dimensions['B'].width=18;ws.column_dimensions['C'].width=12;ws.column_dimensions['D'].width=45

r=1;c=ws.cell(row=r,column=1,value="VISIT VIETNAM — MÔ HÌNH TÀI CHÍNH v4 — GIẢ ĐỊNH");c.font=Font(bold=True,size=14,color="1F4E79")
ws.merge_cells("A1:D1");r+=1
ws.cell(row=r,column=1,value="v4.0 | Ô vàng = giá trị có thể chỉnh | Chữ xanh = placeholder | Named ranges trong mọi công thức").font=Font(italic=True,size=10,color="666666");r+=2

sections=[
    ("THỜI GIAN",[
        ("base_year","Năm cơ sở",2025,"Năm","Năm bắt đầu dự án",FMT_INT),
        ("start_month","Tháng bắt đầu",7,"Tháng","Tháng 7",FMT_INT),
    ]),
    ("DỮ LIỆU THỊ TRƯỜNG",[
        ("TAM_enterprises","Tổng doanh nghiệp du lịch VN",40000,"","GSO/VNAT",FMT_INT),
        ("SAM_enterprises","DN có thể phục vụ",15000,"","Hoạt động số",FMT_INT),
        ("ecom_market_base","Thị trường TMĐT du lịch (năm gốc)",82500,"Tỷ VND","e-Conomy SEA",FMT_MN),
        ("ecom_growth_rate","Tăng trưởng TMĐT du lịch/năm",0.08,"%","Dữ liệu lịch sử",FMT_PCT),
    ]),
    ("ĐƯỜNG CONG TĂNG TRƯỞNG (S-CURVE)",[
        ("adoption_max","Trần doanh nghiệp (tối đa)",1200,"","Tiệm cận S-curve",FMT_INT),
        ("adoption_midpoint","Năm tăng trưởng nhanh nhất",2028,"Năm","Điểm uốn",FMT_INT),
        ("adoption_steepness","Hệ số dốc",0.9,"","Cao hơn = dốc hơn",FMT_DEC),
        ("pilot_count","DN ban đầu (thí điểm)",15,"","Đối tác thí điểm",FMT_INT),
        ("churn_rate","Tỷ lệ không gia hạn/năm",0.20,"%","Churn chứng nhận",FMT_PCT),
    ]),
    ("GIÁ DỊCH VỤ",[
        ("cert_fee_yr1","Phí chứng nhận/DN/năm",25,"Tr VND","Gói thuê bao năm",FMT_INT),
        ("cert_fee_escalation","Tăng phí chứng nhận/năm",0.05,"%","Tăng giá",FMT_PCT),
        ("tx_take_rate","Phí giao dịch (% GMV)",0.02,"%","Hoa hồng nền tảng",FMT_PCT),
        ("platform_market_share","% GMV TMĐT du lịch qua VV",0.005,"%","Mục tiêu thực tế",FMT_PCT),
        ("daas_price","Giá DaaS/triệu VCR",30,"Tr VND","Thuê bao dữ liệu",FMT_INT),
        ("ad_cpm","CPM quảng cáo",150,"K VND","Benchmark B2B",FMT_INT),
        ("ad_impressions_base","Lượt hiển thị QC/tháng (K)",500,"K","Tỷ lệ theo DN",FMT_INT),
        ("training_fee","Phí đào tạo/phiên",15,"Tr VND","Nguồn thu MỚI",FMT_INT),
        ("audit_fee","Phí kiểm toán chứng nhận",10,"Tr VND","Nguồn thu MỚI",FMT_INT),
        ("training_pct","% DN mua đào tạo",0.30,"%","Ước tính",FMT_PCT),
    ]),
    ("THAM SỐ CHI PHÍ",[
        ("capex_maintenance_pct","Bảo trì (% CAPEX/năm)",0.18,"%","Chuẩn ngành 15-20%",FMT_PCT),
        ("opex_marketing","Marketing (cơ sở/năm)",7000,"Tr VND","Dữ liệu gốc",FMT_MN),
        ("marketing_growth","Tăng trưởng marketing/năm",0.15,"%","Tăng theo tham vọng",FMT_PCT),
        ("opex_cloud_base","Cloud cơ sở (năm)",3250,"Tr VND","Dữ liệu gốc",FMT_MN),
        ("cloud_per_1M_vcr","Chi phí cloud/triệu VCR",8,"Tr VND","Chi phí biến đổi",FMT_INT),
        ("vcr_per_enterprise","VCR/DN/tháng",10000,"bản ghi","Ước tính lưu lượng",FMT_INT),
        ("opex_external","Đối ngoại (năm)",400,"Tr VND","Dữ liệu gốc",FMT_MN),
        ("opex_legal","Pháp lý & tuân thủ (năm)",240,"Tr VND","Dữ liệu gốc",FMT_MN),
        ("opex_branding","Thương hiệu & sự kiện (năm)",4000,"Tr VND","Dữ liệu gốc",FMT_MN),
        ("opex_security","Bảo mật/pentest (năm)",250,"Tr VND","Dữ liệu gốc",FMT_MN),
        ("opex_pilot_eval","Đánh giá thí điểm (chỉ Y1)",200,"Tr VND","Chi phí 1 lần",FMT_MN),
        ("payment_proc_pct","Phí cổng thanh toán",0.025,"%","Giá vốn giao dịch",FMT_PCT),
        ("opex_support_per_ent","Hỗ trợ KH/DN/năm",2,"Tr VND","Tỷ lệ theo số DN",FMT_INT),
        ("opex_partner_mgmt","Quản lý đối tác (năm)",500,"Tr VND","Quan hệ đối tác",FMT_MN),
        ("opex_contingency_pct","Dự phòng (% tổng OPEX)",0.10,"%","Đệm rủi ro",FMT_PCT),
    ]),
    ("THAM SỐ NHÂN SỰ",[
        ("salary_mgmt","Lương TB/tháng — Quản lý",45,"Tr VND","GĐ/PM",FMT_INT),
        ("salary_dev","Lương TB/tháng — Phát triển",35,"Tr VND","Kỹ sư",FMT_INT),
        ("salary_bizops","Lương TB/tháng — Vận hành KD",20,"Tr VND","BD/hỗ trợ",FMT_INT),
        ("salary_data","Lương TB/tháng — Dữ liệu",40,"Tr VND","Data eng",FMT_INT),
        ("salary_escalation","Tăng lương hàng năm",0.10,"%","Thị trường VN",FMT_PCT),
        ("bhxh_rate","Bảo hiểm XH (phần DN)",0.215,"%","BHXH+BHYT+BHTN",FMT_PCT),
        ("recruit_cost_pct","Phí tuyển dụng (% lương Y1)",0.20,"%","Headhunter",FMT_PCT),
    ]),
    ("THAM SỐ TÀI CHÍNH",[
        ("cit_rate","Thuế TNDN",0.20,"%","Luật VN",FMT_PCT),
        ("discount_rate","Tỷ suất chiết khấu (WACC)",0.15,"%","Điều chỉnh rủi ro",FMT_PCT),
        ("depreciation_years","Thời gian khấu hao",5,"Năm","Đường thẳng",FMT_INT),
        ("ar_days","Số ngày phải thu",45,"Ngày","Kỳ thu tiền",FMT_INT),
        ("ap_days","Số ngày phải trả",30,"Ngày","Kỳ thanh toán",FMT_INT),
        ("stage2_gate","Ngưỡng Giai đoạn 2 (DN tối thiểu)",300,"","Kích hoạt phí GD",FMT_INT),
    ]),
]

for sec_title, params in sections:
    sc(ws,r,1,sec_title)
    for col in [2,3,4]:ws.cell(row=r,column=col).fill=SC_FILL;ws.cell(row=r,column=col).border=BD
    r+=1
    hd(ws,r,1,"Tham số");hd(ws,r,2,"Giá trị");hd(ws,r,3,"Đơn vị");hd(ws,r,4,"Ghi chú");r+=1
    for nm,lbl,val,unit,note,fmt in params:
        lb(ws,r,1,lbl);ph(ws,r,2,val,fmt)
        ws.cell(row=r,column=3,value=unit).font=LB;ws.cell(row=r,column=3).border=BD
        ws.cell(row=r,column=4,value=note).font=NOTE;ws.cell(row=r,column=4).border=BD
        add_nm(nm,r);r+=1
    r+=1

# ===== ĐẦU TƯ BAN ĐẦU =====
ws=wb.create_sheet("Đầu tư ban đầu");ws.sheet_properties.tabColor="C00000"
ws.column_dimensions['A'].width=40;ws.column_dimensions['B'].width=18;ws.column_dimensions['C'].width=12;ws.column_dimensions['D'].width=40
r=1;sc(ws,r,1,"ĐẦU TƯ BAN ĐẦU — CHI TIẾT CAPEX (Tr VND)");r+=2
hd(ws,r,1,"Hạng mục CAPEX");hd(ws,r,2,"Số tiền (Tr VND)");hd(ws,r,3,"Thời điểm");hd(ws,r,4,"Ghi chú");r+=1

items=[
    ("capex_cloud_setup","Hạ tầng Cloud",800,"H2 2025","Setup AWS/GCP ban đầu"),
    ("capex_backend","Phát triển Backend",1200,"H2 2025","API, DB, logic lõi"),
    ("capex_frontend","Frontend + mobile",900,"H2 2025","Web + ứng dụng di động"),
    ("capex_data_platform","Nền tảng dữ liệu",600,"H2 2025","Pipeline, phân tích"),
    ("capex_security_setup","Hạ tầng bảo mật",250,"H2 2025","Pentest, xác thực"),
    ("capex_ux_design","Thiết kế UX/UI",350,"H2 2025","Design system"),
    ("capex_integration","Tích hợp bên thứ 3",400,"H2 2025","Cổng TT, API OTA"),
    ("capex_devops","DevOps & CI/CD",280,"H2 2025","Pipeline triển khai"),
    ("capex_testing","QA & kiểm thử",200,"H2 2025","Kiểm thử tự động"),
    ("capex_contingency","Dự phòng CAPEX (10%)",300,"H2 2025","Đệm phát sinh"),
]
first=r
for nm2,lbl,val,timing,note in items:
    lb(ws,r,1,lbl);ph(ws,r,2,val,FMT)
    ws.cell(row=r,column=3,value=timing).font=LB;ws.cell(row=r,column=3).border=BD
    ws.cell(row=r,column=4,value=note).font=NOTE;ws.cell(row=r,column=4).border=BD
    dn=DefinedName(nm2,attr_text=f"'Đầu tư ban đầu'!$B${r}");wb.defined_names.add(dn)
    r+=1
last=r-1;r+=1
lb(ws,r,1,"TỔNG ĐẦU TƯ BAN ĐẦU")
tot(ws,r,2,f"=SUM(B{first}:B{last})",FMT)
dn=DefinedName("capex_total",attr_text=f"'Đầu tư ban đầu'!$B${r}");wb.defined_names.add(dn)
r+=2

sc(ws,r,1,"LỊCH KHẤU HAO");r+=1
hd(ws,r,1,"Năm");hd(ws,r,2,"Khấu hao/năm");hd(ws,r,3,"Khấu hao lũy kế");hd(ws,r,4,"Giá trị còn lại");r+=1
for i in range(6):
    yr=2025+i;lb(ws,r,1,str(yr))
    if i<5:fm(ws,r,2,"=capex_total/depreciation_years",FMT)
    else:fm(ws,r,2,"=0",FMT)
    if i==0:fm(ws,r,3,f"=B{r}",FMT)
    else:fm(ws,r,3,f"=C{r-1}+B{r}",FMT)
    fm(ws,r,4,f"=capex_total-C{r}",FMT);r+=1

# ===== KẾ HOẠCH NHÂN SỰ =====
MN=['','Th1','Th2','Th3','Th4','Th5','Th6','Th7','Th8','Th9','Th10','Th11','Th12']
TC=[]
for m in range(7,13):TC.append((f"{MN[m]} 2025",2025,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2026",2026,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2027",2027,'m',m))
for yr in [2028,2029,2030]:
    for q in range(1,5):TC.append((f"Q{q} {yr}",yr,'q',q))
TSC=3;NTC=42

ws=wb.create_sheet("Nhân sự");ws.sheet_properties.tabColor="7030A0"
ws.column_dimensions['A'].width=30;ws.column_dimensions['B'].width=14
r=1;sc(ws,r,1,"KẾ HOẠCH NHÂN SỰ — BIÊN CHẾ & CHI PHÍ");r+=2
hd(ws,r,1,"Bộ phận / Hạng mục");hd(ws,r,2,"ĐVT")
for i,(l,y,t,m) in enumerate(TC):col=TSC+i;hd(ws,r,col,l);ws.column_dimensions[get_column_letter(col)].width=11
r+=1

sc(ws,r,1,"SỐ LƯỢNG NHÂN SỰ")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1

depts=[
    ("Quản lý","salary_mgmt",[2]*6+[2]*12+[3]*12+[3]*4+[4]*4+[4]*4),
    ("Phát triển","salary_dev",[12]*6+[12]*12+[14]*12+[16]*4+[18]*4+[20]*4),
    ("Vận hành KD","salary_bizops",[2]*6+[6]*12+[10]*12+[18]*4+[28]*4+[35]*4),
    ("Dữ liệu & Phân tích","salary_data",[0]*6+[2]*12+[3]*12+[4]*4+[5]*4+[6]*4),
]
dept_rows={}
for dn,sal,counts in depts:
    lb(ws,r,1,dn);ws.cell(row=r,column=2,value="người").font=LB
    for i in range(NTC):ph(ws,r,TSC+i,counts[i],'#,##0')
    dept_rows[dn]=r;r+=1

hc_tot_r=r
lb(ws,r,1,"TỔNG NHÂN SỰ");ws.cell(row=r,column=2,value="người").font=TF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    tot(ws,r,TSC+i,f"={cl}{dept_rows['Quản lý']}+{cl}{dept_rows['Phát triển']}+{cl}{dept_rows['Vận hành KD']}+{cl}{dept_rows['Dữ liệu & Phân tích']}",FMT)
r+=2

sc(ws,r,1,"TUYỂN MỚI")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
new_h_r=r
lb(ws,r,1,"Tuyển mới");ws.cell(row=r,column=2,value="người").font=LB
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    if i==0:fm(ws,r,TSC+i,f"={cl}{hc_tot_r}",FMT)
    else:fm(ws,r,TSC+i,f"=MAX(0,{cl}{hc_tot_r}-{get_column_letter(TSC+i-1)}{hc_tot_r})",FMT)
r+=2

sc(ws,r,1,"CHI PHÍ LƯƠNG (Tr VND)")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
sal_rows={}
for dn,sal,_ in depts:
    lb(ws,r,1,f"CP {dn}");ws.cell(row=r,column=2,value="Tr VND").font=LB
    for i,(l,yr,t,m) in enumerate(TC):
        col=TSC+i;cl=get_column_letter(col);ye=yr-2025;mult=1 if t=='m' else 3
        fm(ws,r,col,f"={cl}{dept_rows[dn]}*{sal}*{mult}*(1+salary_escalation)^{ye}*(1+bhxh_rate)",FMT)
    sal_rows[dn]=r;r+=1

sal_tot_r=r
lb(ws,r,1,"TỔNG LƯƠNG");ws.cell(row=r,column=2,value="Tr VND").font=TF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    tot(ws,r,TSC+i,f"={cl}{sal_rows['Quản lý']}+{cl}{sal_rows['Phát triển']}+{cl}{sal_rows['Vận hành KD']}+{cl}{sal_rows['Dữ liệu & Phân tích']}",FMT)
r+=1
rec_r=r
lb(ws,r,1,"Tuyển dụng");ws.cell(row=r,column=2,value="Tr VND").font=LB
for i in range(NTC):cl=get_column_letter(TSC+i);fm(ws,r,TSC+i,f"={cl}{new_h_r}*30*12*recruit_cost_pct",FMT)
r+=1
hr_tot_r=r
lb(ws,r,1,"TỔNG CHI PHÍ NHÂN SỰ");ws.cell(row=r,column=2,value="Tr VND").font=TF
for i in range(NTC):cl=get_column_letter(TSC+i);tot(ws,r,TSC+i,f"={cl}{sal_tot_r}+{cl}{rec_r}",FMT)

state={'named':named,'out':OUT,'final':FINAL,'hr_tot_r':hr_tot_r,'new_h_r':new_h_r,
    'hc_tot_r':hc_tot_r,'dept_rows':{k:v for k,v in dept_rows.items()}}
with open('/tmp/v4vi_state.json','w') as f:json.dump(state,f)
wb.save(OUT)
print(f"Part 1 done: Assumptions + Đầu tư + Nhân sự")
print(f"Named ranges: {len(named)}")
