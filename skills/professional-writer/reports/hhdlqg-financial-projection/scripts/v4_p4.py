"""
v4 Part 4: Scenarios + Budget 2025/2026 + Guide + finalize
"""
import openpyxl, json, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/v4_state.json') as f: state=json.load(f)
OUT=state['out'];FINAL=state['final']

HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LBF=Font(size=11);TF=Font(bold=True,size=11)
TFL=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
PH_FILL=PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid")
PH_FONT=Font(color="0000FF",size=11)
BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FMT='#,##0';FP='0.0%'

def hd(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def sc(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
def lb(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LBF;cl.border=BD
def fm(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=Font(size=11);cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def tot(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=TF;cl.fill=TFL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def ph(ws,r,c,v,fmt=None):cl=ws.cell(row=r,column=c,value=v);cl.font=PH_FONT;cl.fill=PH_FILL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)

MN=['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

wb=openpyxl.load_workbook(OUT)

# ===== SCENARIOS =====
ws=wb.create_sheet("Scenarios");ws.sheet_properties.tabColor="FFC000"
ws.column_dimensions['A'].width=32
for c in ['B','C','D','E','F']:ws.column_dimensions[c].width=16
r=1;sc(ws,r,1,"SCENARIO ANALYSIS");ws.merge_cells("A1:F1");r+=1
ws.cell(row=r,column=1,value="Copy values from a column into Assumptions.").font=Font(italic=True,size=10,color="666666");r+=2
for ci,h in[(1,"Parameter"),(2,"WORST"),(3,"BEAR"),(4,"BASE"),(5,"BULL"),(6,"BEST")]:hd(ws,r,ci,h)
r+=1
scen=[
    ("adoption_max",200,600,"=adoption_max",2000,3500),
    ("adoption_steepness",0.5,0.7,"=adoption_steepness",1.2,1.5),
    ("churn_rate",0.35,0.25,"=churn_rate",0.15,0.10),
    ("platform_market_share",0.001,0.003,"=platform_market_share",0.01,0.02),
    ("cert_fee_yr1",15,20,"=cert_fee_yr1",30,40),
    ("cert_fee_escalation",0.0,0.03,"=cert_fee_escalation",0.08,0.12),
    ("tx_take_rate",0.01,0.015,"=tx_take_rate",0.025,0.03),
    ("opex_marketing",8000,7500,"=opex_marketing",6000,5000),
    ("discount_rate",0.20,0.18,"=discount_rate",0.12,0.10),
    ("stage2_gate",500,400,"=stage2_gate",200,150),
]
for label,w,be,ba,bu,best in scen:
    lb(ws,r,1,label)
    ip=isinstance(w,float) and w<1
    f=FP if ip else FMT
    for ci,v,clr in[(2,w,"CC0000"),(3,be,None),(4,ba,None),(5,bu,None),(6,best,"00B050")]:
        c=ws.cell(row=r,column=ci,value=v);c.border=BD;c.number_format=f
        if clr:c.font=Font(size=11,color=clr)
        if ci==4:c.fill=TFL;c.font=Font(bold=True,size=11)
    r+=1

# ===== BUDGET BUILDER =====
def build_budget(wb,year,months,title):
    ws=wb.create_sheet(title);ws.sheet_properties.tabColor="4472C4"
    ws.column_dimensions['A'].width=35;ws.column_dimensions['B'].width=12
    nm=len(months);tc=3+nm
    r=1;sc(ws,r,1,f"BUDGET {year} (M VND)");r+=2
    hd(ws,r,1,"Item");hd(ws,r,2,"Unit")
    for j,m in enumerate(months):
        col=3+j;hd(ws,r,col,MN[m]);ws.column_dimensions[get_column_letter(col)].width=12
    hd(ws,r,tc,"TOTAL");ws.column_dimensions[get_column_letter(tc)].width=14;r+=1

    def rt(r_):return f"=SUM({get_column_letter(3)}{r_}:{get_column_letter(3+nm-1)}{r_})"
    ye=year-2025

    # Revenue
    sc(ws,r,1,"REVENUE");
    for j in range(nm):ws.cell(row=r,column=3+j).fill=SC_FILL
    ws.cell(row=r,column=tc).fill=SC_FILL;r+=1
    cert_r=r;lb(ws,r,1,"  Cert fees");lb(ws,r,2,"M VND")
    for j,m in enumerate(months):
        yf=year+(m-0.5)/12
        fm(ws,r,3+j,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({yf:.3f}-adoption_midpoint))))*cert_fee_yr1*(1+cert_fee_escalation)^{ye}/12",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    tr_r=r;lb(ws,r,1,"  Training");lb(ws,r,2,"M VND")
    for j,m in enumerate(months):
        yf=year+(m-0.5)/12
        fm(ws,r,3+j,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({yf:.3f}-adoption_midpoint))))*training_pct*training_fee/12",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    au_r=r;lb(ws,r,1,"  Audit");lb(ws,r,2,"M VND")
    for j,m in enumerate(months):
        yf=year+(m-0.5)/12
        fm(ws,r,3+j,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({yf:.3f}-adoption_midpoint))))*audit_fee/12",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    rev_r=r;lb(ws,r,1,"TOTAL REVENUE")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"=SUM({cl}{cert_r}:{cl}{au_r})",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=2

    # CAPEX
    sc(ws,r,1,"CAPEX");
    for j in range(nm):ws.cell(row=r,column=3+j).fill=SC_FILL
    ws.cell(row=r,column=tc).fill=SC_FILL;r+=1
    cx_start=r
    if year==2025:
        capex_items=["capex_cloud_setup","capex_backend","capex_frontend","capex_data_platform","capex_security_setup","capex_ux_design","capex_integration","capex_devops","capex_testing","capex_contingency"]
        capex_labels=["Cloud setup","Backend dev","Frontend+mobile","Data platform","Security","UX/UI design","Integrations","DevOps","QA testing","Contingency"]
        for nm2,lbl in zip(capex_items,capex_labels):
            lb(ws,r,1,f"  {lbl}");lb(ws,r,2,"M VND")
            for j in range(nm):fm(ws,r,3+j,f"={nm2}/{nm}",FMT)
            tot(ws,r,tc,rt(r),FMT);r+=1
    else:
        lb(ws,r,1,"  Maintenance");lb(ws,r,2,"M VND")
        for j in range(nm):fm(ws,r,3+j,"=capex_total*capex_maintenance_pct/12",FMT)
        tot(ws,r,tc,rt(r),FMT);r+=1
    cx_end=r-1
    cx_r=r;lb(ws,r,1,"TOTAL CAPEX")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"=SUM({cl}{cx_start}:{cl}{cx_end})",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=2

    # OPEX
    sc(ws,r,1,"OPEX");
    for j in range(nm):ws.cell(row=r,column=3+j).fill=SC_FILL
    ws.cell(row=r,column=tc).fill=SC_FILL;r+=1
    ox_start=r
    hc={2025:{"Mgmt(2)":"salary_mgmt*2","Dev(12)":"salary_dev*12","BizOps(2)":"salary_bizops*2"},
        2026:{"Mgmt(2)":"salary_mgmt*2","Dev(12)":"salary_dev*12","BizOps(6)":"salary_bizops*6","Data(2)":"salary_data*2"}}
    for dept,formula in hc.get(year,hc[2026]).items():
        lb(ws,r,1,f"  HR — {dept}");lb(ws,r,2,"M VND")
        for j in range(nm):fm(ws,r,3+j,f"={formula}*(1+salary_escalation)^{ye}*(1+bhxh_rate)",FMT)
        tot(ws,r,tc,rt(r),FMT);r+=1

    # Recruitment
    lb(ws,r,1,"  Recruitment");lb(ws,r,2,"M VND")
    new_h=16 if year==2025 else 6
    for j in range(nm):
        if j==0:fm(ws,r,3+j,f"={new_h}*30*12*recruit_cost_pct",FMT)
        else:fm(ws,r,3+j,"=0",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1

    opex_lines=[
        ("Marketing",f"=opex_marketing*(1+marketing_growth)^{ye}/12"),
        ("Branding","=opex_branding/12"),("Cloud","=opex_cloud_base/12"),
        ("Security","=opex_security/12"),("External","=opex_external/12"),
        ("Legal","=opex_legal/12"),("Partner mgmt","=opex_partner_mgmt/12"),
    ]
    if year==2025:opex_lines.append(("Pilot eval","=opex_pilot_eval/6"))
    for lbl,formula in opex_lines:
        lb(ws,r,1,f"  {lbl}");lb(ws,r,2,"M VND")
        for j in range(nm):fm(ws,r,3+j,formula,FMT)
        tot(ws,r,tc,rt(r),FMT);r+=1

    # Support
    lb(ws,r,1,"  Customer support");lb(ws,r,2,"M VND")
    for j,m in enumerate(months):
        yf=year+(m-0.5)/12
        fm(ws,r,3+j,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({yf:.3f}-adoption_midpoint))))*opex_support_per_ent/12",FMT)
    tot(ws,r,tc,rt(r),FMT);ox_end=r;r+=1

    ox_sub=r;lb(ws,r,1,"OPEX Subtotal")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"=SUM({cl}{ox_start}:{cl}{ox_end})",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    cg=r;lb(ws,r,1,"  Contingency");lb(ws,r,2,"M VND")
    for j in range(nm):cl=get_column_letter(3+j);fm(ws,r,3+j,f"={cl}{ox_sub}*opex_contingency_pct",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    ox_tot=r;lb(ws,r,1,"TOTAL OPEX")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"={cl}{ox_sub}+{cl}{cg}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=2

    # Summary
    sc(ws,r,1,"SUMMARY");
    for j in range(nm):ws.cell(row=r,column=3+j).fill=SC_FILL
    ws.cell(row=r,column=tc).fill=SC_FILL;r+=1
    sr=r;lb(ws,r,1,"Revenue")
    for j in range(nm):cl=get_column_letter(3+j);fm(ws,r,3+j,f"={cl}{rev_r}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    lb(ws,r,1,"CAPEX")
    for j in range(nm):cl=get_column_letter(3+j);fm(ws,r,3+j,f"={cl}{cx_r}",FMT)
    tot(ws,r,tc,rt(r),FMT);scx=r;r+=1
    lb(ws,r,1,"OPEX")
    for j in range(nm):cl=get_column_letter(3+j);fm(ws,r,3+j,f"={cl}{ox_tot}",FMT)
    tot(ws,r,tc,rt(r),FMT);sox=r;r+=1
    tb=r;lb(ws,r,1,"TOTAL BUDGET")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"={cl}{scx}+{cl}{sox}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    nr=r;lb(ws,r,1,"NET (Rev-Budget)")
    for j in range(nm):cl=get_column_letter(3+j);tot(ws,r,3+j,f"={cl}{sr}-{cl}{tb}",FMT)
    tot(ws,r,tc,rt(r),FMT);r+=1
    lb(ws,r,1,"Cumulative")
    for j in range(nm):
        cl=get_column_letter(3+j)
        if j==0:fm(ws,r,3+j,f"={cl}{nr}",FMT)
        else:fm(ws,r,3+j,f"={get_column_letter(3+j-1)}{r}+{cl}{nr}",FMT)
    tot(ws,r,tc,f"={get_column_letter(3+nm-1)}{r}",FMT)

# Build budgets
build_budget(wb,2025,list(range(7,13)),"Budget 2025")
build_budget(wb,2026,list(range(1,13)),"Budget 2026")

# ===== GUIDE =====
ws=wb.create_sheet("Guide",0)
ws.sheet_properties.tabColor="00B050"
ws.column_dimensions['A'].width=25;ws.column_dimensions['B'].width=55;ws.column_dimensions['C'].width=35
r=1;ws.cell(row=r,column=1,value="VISIT VIETNAM — FINANCIAL MODEL v4 — GUIDE").font=Font(bold=True,size=16,color="1F4E79")
ws.merge_cells("A1:C1");r+=1
ws.cell(row=r,column=1,value="v4.0 | 2026-03-21 | Fixed named ranges (v3 had broken refs from row insertion)").font=Font(italic=True,size=10,color="666666");r+=2

sc(ws,r,1,"SHEET DIRECTORY");r+=1
hd(ws,r,1,"Sheet");hd(ws,r,2,"Purpose");hd(ws,r,3,"Notes");r+=1
sheets=[
    ("Guide","Model documentation, named ranges list, formula patterns, instructions","Read first"),
    ("Assumptions","49 named ranges. ALL inputs live here (yellow cells).","Only edit yellow cells"),
    ("Initial Investment","10 CAPEX items → capex_total. Depreciation schedule.","capex_total used everywhere"),
    ("HR Plan","4 depts × 42 periods. Salary = headcount × named_salary × (1+escalation)^yrs × (1+bhxh)","Yellow = editable headcount"),
    ("Market Sizing","S-curve adoption. Year fracs from base_year+start_month. VCR projections.","No hardcoded years"),
    ("Revenue Schedule","6 streams. Stage 2 gated on stage2_gate.","All named range formulas"),
    ("Cost Schedule","CAPEX (Initial Investment) + OPEX (HR Plan + 10 categories + contingency)","Cross-sheet references"),
    ("P&L","Revenue → COGS → GP → OPEX → EBITDA → D&A → EBIT → CIT → NI + margins","capex_total/depreciation_years"),
    ("Cash Flow","OCF + ICF → FCF + Cumulative. Working capital via ar_days.","All named ranges"),
    ("DCF & Returns","Annual FCF, NPV(discount_rate,...), IRR, ROI, max cash burn","Single NPV formula"),
    ("Scenarios","5 scenarios: Worst/Bear/Base/Bull/Best. Copy column → Assumptions.","Shows why v2 had 214% IRR"),
    ("Budget 2025","Monthly Jul-Dec. Revenue + CAPEX (10 items) + OPEX detail + contingency.","Approval-ready"),
    ("Budget 2026","Monthly Jan-Dec. Same structure, HR scales to 22 FTE.","Approval-ready"),
]
for n,p,nt in sheets:lb(ws,r,1,n);lb(ws,r,2,p);lb(ws,r,3,nt);r+=1
r+=1

sc(ws,r,1,"NAMED RANGES (49 total)");r+=1
hd(ws,r,1,"Name");hd(ws,r,2,"Description");hd(ws,r,3,"Example formula usage");r+=1
nr_doc=[
    ("base_year","Project start year","Year frac = base_year + (start_month-1+idx)/12"),
    ("adoption_max","S-curve ceiling","MAX(pilot_count, adoption_max/(1+EXP(...)))"),
    ("cert_fee_yr1","Annual cert fee (M VND)","net_ent × cert_fee_yr1 × (1+escalation)^yrs"),
    ("capex_total","Total CAPEX (from Initial Investment)","capex_total/depreciation_years for D&A"),
    ("depreciation_years","Depreciation period","capex_total/depreciation_years/period_divisor"),
    ("salary_bizops","BizOps monthly salary","headcount × salary_bizops × months × (1+escalation)^yrs × (1+bhxh)"),
    ("bhxh_rate","Social insurance rate","Multiplied on all salary costs"),
    ("cit_rate","Corporate income tax","IF(EBIT>0, -EBIT×cit_rate, 0)"),
    ("discount_rate","WACC","NPV(discount_rate, FCF_array)"),
    ("stage2_gate","Min enterprises for Stage 2","IF(net_ent >= stage2_gate, tx_rev, 0)"),
]
for n,d,u in nr_doc:lb(ws,r,1,n);lb(ws,r,2,d);lb(ws,r,3,u);r+=1
r+=1

sc(ws,r,1,"v3→v4 FIX");r+=1
for line in[
    "v3 had broken named ranges because inserting rows shifted cells but DefinedName refs didn't update.",
    "v4 is rebuilt from scratch — all 49 named ranges verified pointing to correct cells.",
    "All formulas verified: capex_total/depreciation_years, salary_bizops*(1+salary_escalation), etc."
]:ws.cell(row=r,column=1,value=line).font=Font(size=10,color="444444");ws.merge_cells(f"A{r}:C{r}");r+=1

# Reorder
order=["Guide","Assumptions","Initial Investment","HR Plan","Market Sizing","Revenue Schedule","Cost Schedule","P&L","Cash Flow","DCF & Returns","Scenarios","Budget 2025","Budget 2026"]
for i,n in enumerate(order):
    if n in wb.sheetnames:wb.move_sheet(n,offset=i-wb.sheetnames.index(n))

wb.save(OUT)
print(f"v4 complete: {OUT}")
print(f"Sheets: {wb.sheetnames}")

# Copy to final
try:
    shutil.copy2(OUT,FINAL);print(f"Copied to: {FINAL}")
except Exception as e:print(f"Copy failed (close Excel first): {e}\nFile at: {OUT}")
