"""
Add glossary / terms explanation to Guide sheet in v4.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

SRC = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\hhdlqg-financial-projection\VisitVietnam_FinancialModel_v4.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb['Guide']

SC_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SC_FONT = Font(bold=True, size=11, color="1F4E79")
HD_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HD_FONT = Font(color="FFFFFF", bold=True, size=11)
LBF = Font(size=11)
BD = Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),
    top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))

def hd(r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def lb(r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LBF;cl.border=BD
def sc(r,v):
    cl=ws.cell(row=r,column=1,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
    for col in [2,3]:ws.cell(row=r,column=col).fill=SC_FILL;ws.cell(row=r,column=col).border=BD

r = ws.max_row + 3

sc(r, "GLOSSARY — KEY TERMS & ABBREVIATIONS"); r += 1
hd(r,1,"Term"); hd(r,2,"Full Name / Meaning"); hd(r,3,"In this model"); r += 1

terms = [
    # Financial terms
    ("NPV", "Net Present Value", "Sum of all future cash flows discounted to today. Positive = project creates value."),
    ("IRR", "Internal Rate of Return", "Discount rate at which NPV = 0. Higher = better return. Compare vs WACC."),
    ("WACC", "Weighted Average Cost of Capital", "Minimum return investors expect. Used as discount_rate (15% Base)."),
    ("ROI", "Return on Investment", "(Cumulative CF + Investment) / Investment. How much you get back per VND invested."),
    ("EBITDA", "Earnings Before Interest, Tax, Depreciation & Amortization", "Operating profit before non-cash charges and tax. Key profitability metric."),
    ("EBIT", "Earnings Before Interest & Tax", "= EBITDA - Depreciation. Pre-tax profit."),
    ("CIT", "Corporate Income Tax", "Vietnamese standard rate = 20%. Applied only when EBIT > 0."),
    ("FCF", "Free Cash Flow", "= Operating CF + Investing CF. Cash actually available after all spending."),
    ("D&A", "Depreciation & Amortization", "Non-cash expense spreading CAPEX over useful life (5 years here)."),
    ("NOPAT", "Net Operating Profit After Tax", "= EBIT × (1 - CIT rate). What the business earns after tax."),
    # Cost terms
    ("CAPEX", "Capital Expenditure", "One-time investment in assets (platform dev 5,280M). Depreciated over time."),
    ("OPEX", "Operating Expenditure", "Ongoing costs: HR, marketing, cloud, legal, etc. Expensed immediately."),
    ("COGS", "Cost of Goods Sold", "Direct costs tied to revenue: payment processing fees (2.5% of tx revenue)."),
    ("BHXH", "Bảo hiểm xã hội (Social Insurance)", "Employer pays ~21.5%: BHXH 17.5% + BHYT 3% + BHTN 1%."),
    # Market terms
    ("TAM", "Total Addressable Market", "All tourism enterprises in Vietnam (~40,000)."),
    ("SAM", "Serviceable Addressable Market", "Digitally active, certifiable subset (~15,000)."),
    ("SOM", "Serviceable Obtainable Market", "What VV can realistically capture (= adoption_max)."),
    ("GMV", "Gross Merchandise Value", "Total value of transactions flowing through the platform."),
    ("VCR", "Visitor Certified Records", "Data records generated per enterprise. Basis for DaaS pricing."),
    # Model terms
    ("S-curve", "Logistic Growth Curve", "Adoption follows slow→fast→slow pattern. Formula: max/(1+EXP(-k*(t-midpoint)))"),
    ("Stage 1", "Certification Platform (2025-2027)", "Revenue from cert fees, training, audits. Doesn't need transaction volume."),
    ("Stage 2", "Transaction Platform (2028-2030)", "Revenue from tx fees, DaaS, ads. Gated: needs ≥300 enterprises (stage2_gate)."),
    ("Gate", "Stage 2 activation threshold", "= stage2_gate (300 enterprises). IF net_ent ≥ gate → Stage 2 revenue activates."),
    ("Churn", "Non-renewal rate", "% of enterprises that leave each year. Reduces net_enterprises."),
    ("CAC", "Customer Acquisition Cost", "Marketing spend / new enterprises acquired. Not explicitly modeled."),
    ("LTV", "Lifetime Value", "Total revenue from one enterprise over its lifetime. = ARPU / churn_rate."),
    ("ARPU", "Average Revenue Per User", "Total revenue / net_enterprises. Annual revenue per certified enterprise."),
    # Excel terms
    ("Named Range", "Excel defined name for a cell", "e.g. cert_fee_yr1 → Assumptions!$B$26. Used in formulas for clarity."),
    ("Placeholder", "Yellow cell with blue font", "Editable input value. Change it → entire model recalculates."),
]

for term, full, context in terms:
    lb(r, 1, term); lb(r, 2, full); lb(r, 3, context)
    r += 1

wb.save(SRC)
print(f"Guide updated with {len(terms)} glossary terms")
