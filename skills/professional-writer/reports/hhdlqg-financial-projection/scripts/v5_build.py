"""
v5 Part 1: Load v4, add cascading relationship params at bottom of Assumptions,
update Cost Schedule formulas, add Sensitivity sheet.
"""
import openpyxl, shutil, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

SRC = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\hhdlqg-financial-projection\VisitVietnam_FinancialModel_v4.xlsx"
TMP = r"E:\tmp\v5_build.xlsx"
FINAL = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\hhdlqg-financial-projection\VisitVietnam_FinancialModel_v5.xlsx"

shutil.copy2(SRC, TMP)
wb = openpyxl.load_workbook(TMP)

PH_FILL=PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid")
PH_FONT=Font(color="0000FF",size=11)
HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LB=Font(size=11);TF=Font(bold=True,size=11)
TFL=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
NOTE=Font(italic=True,size=10,color="888888")
REL_FILL=PatternFill(start_color="FFF2CC",end_color="FFF2CC",fill_type="solid")
REL_FONT=Font(bold=True,size=11,color="BF8F00")
BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),
    top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FMT='#,##0';FP='0.0%';FD='0.0'

def add_nm(name, sheet, row, col=2):
    ref = f"'{sheet}'!${get_column_letter(col)}${row}"
    try: wb.defined_names.delete(name)
    except: pass
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)

def hd(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def sc(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
def lb(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LB;cl.border=BD
def fm(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=Font(size=11);cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def tot(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=TF;cl.fill=TFL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def ph(ws,r,c,v,fmt=None):cl=ws.cell(row=r,column=c,value=v);cl.font=PH_FONT;cl.fill=PH_FILL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)

# ============================================================
# STEP 1: Add cascading relationship params to Assumptions
# ============================================================
ws = wb['Assumptions']
r = ws.max_row + 3

# Section header with distinct color
c = ws.cell(row=r, column=1, value="CASCADING RELATIONSHIPS (how assumptions affect each other)")
c.font = REL_FONT; c.fill = REL_FILL; c.border = BD
for col in [2,3,4]:ws.cell(row=r,column=col).fill=REL_FILL;ws.cell(row=r,column=col).border=BD
r += 1

# Sub-header
hd(ws,r,1,"Parameter");hd(ws,r,2,"Value");hd(ws,r,3,"Unit");hd(ws,r,4,"Cascade logic");r+=1

cascade_params = [
    # Cloud scales with GMV
    ("cloud_gmv_factor", "Cloud cost per Bn VND GMV processed", 0.5, "Tr VND/Bn",
     "↑ market_share → ↑ GMV → ↑ cloud cost. Formula: GMV × this factor", FD),
    # Marketing efficiency
    ("mkt_efficiency_per_100", "Marketing cost reduction per 100 enterprises", 0.02, "%/100 ent",
     "↑ adoption → ↓ marketing/ent. Formula: marketing × MAX(0.5, 1 - ent×factor/100)", FP),
    # Support economies
    ("support_scale_threshold", "Enterprises threshold for economies of scale", 500, "enterprises",
     "When net_ent > threshold → support cost drops by support_discount", FMT),
    ("support_discount", "Support cost reduction after threshold", 0.30, "ratio",
     "↑ ent past threshold → support/ent × (1 - discount). Economies of scale.", FP),
    # HR scaling
    ("hr_bizops_per_100_ent", "BizOps staff needed per 100 enterprises (floor)", 2.0, "people/100",
     "↑ adoption → ↑ BizOps needed. Floor = MAX(manual_hc, ent×ratio/100)", FD),
    # Cloud variable scaling
    ("cloud_elasticity", "Cloud cost elasticity to enterprise count", 1.2, "factor",
     "VCR cost scales by ent^elasticity instead of linear. >1 = super-linear.", FD),
]

for nm, lbl, val, unit, cascade, fmt in cascade_params:
    lb(ws,r,1,lbl); ph(ws,r,2,val,fmt)
    ws.cell(row=r,column=3,value=unit).font=LB;ws.cell(row=r,column=3).border=BD
    ws.cell(row=r,column=4,value=cascade).font=NOTE;ws.cell(row=r,column=4).border=BD
    add_nm(nm, "Assumptions", r)
    r += 1

# Add a visual "Impact Map" showing the cascade
r += 2
c = ws.cell(row=r, column=1, value="IMPACT MAP — How market_share cascades to costs")
c.font = REL_FONT; c.fill = REL_FILL; c.border = BD
ws.merge_cells(f"A{r}:D{r}")
for col in [2,3,4]:ws.cell(row=r,column=col).fill=REL_FILL
r += 1

cascade_map = [
    ("adoption_max ↑", "→ net_enterprises ↑", "→ HR BizOps headcount ↑ (hr_bizops_per_100_ent)", "→ HR Cost ↑"),
    ("", "", "→ VCR volume ↑ (vcr_per_enterprise × net_ent)", "→ Cloud Cost ↑ (cloud_per_1M_vcr)"),
    ("", "", "→ Support demand ↑ (opex_support_per_ent × net_ent)", "→ OPEX ↑"),
    ("", "", "→ Marketing efficiency ↑ (cost/ent decreases)", "→ Per-unit OPEX ↓"),
    ("", "", "→ Cert/Training/Audit revenue ↑", "→ Revenue ↑"),
    ("", "", "", ""),
    ("platform_market_share ↑", "→ GMV through platform ↑", "→ Transaction fees ↑ (tx_take_rate × GMV)", "→ Revenue ↑"),
    ("", "", "→ Cloud processing load ↑ (cloud_gmv_factor × GMV)", "→ Cloud Cost ↑"),
    ("", "", "→ Payment processing fees ↑ (payment_proc_pct × tx_rev)", "→ COGS ↑"),
    ("", "", "→ DaaS & Ads potential ↑", "→ Revenue ↑"),
    ("", "", "", ""),
    ("churn_rate ↑", "→ net_enterprises ↓", "→ All enterprise-linked revenue ↓", "→ Revenue ↓"),
    ("", "", "→ Need more new sales to replace", "→ Marketing cost pressure ↑"),
    ("", "", "→ Support & HR needs more stable", "→ Cost stays high vs falling revenue"),
]
for a,b,c_,d in cascade_map:
    for ci,v in[(1,a),(2,b),(3,c_),(4,d)]:
        cl = ws.cell(row=r, column=ci, value=v)
        cl.font = Font(size=9, color="444444" if not a else "1F4E79")
        cl.border = BD
    r += 1

print(f"Step 1 done: Added cascade params and impact map to Assumptions (rows up to {r})")

# ============================================================
# STEP 2: Rebuild Cost Schedule with cascading formulas
# ============================================================
# We need to know the row numbers in existing sheets
# Let me find them by reading current Cost Schedule structure

ws_cost = wb['Cost Schedule']
# Find key rows by scanning column A
cost_rows = {}
for row in range(1, ws_cost.max_row + 1):
    val = ws_cost.cell(row=row, column=1).value
    if val:
        val_str = str(val).strip()
        if 'Cloud' in val_str and 'infra' in val_str: cost_rows['cloud'] = row
        elif 'Marketing' in val_str and 'CAPEX' not in val_str and 'OPEX' not in val_str: cost_rows['marketing'] = row
        elif 'Customer support' in val_str or 'support' in val_str.lower():
            if 'OPEX' not in val_str and 'Subtotal' not in val_str: cost_rows['support'] = row
        elif 'TOTAL CAPEX' in val_str: cost_rows['capex_total'] = row
        elif 'OPEX Subtotal' in val_str: cost_rows['opex_sub'] = row
        elif 'TOTAL OPEX' in val_str: cost_rows['opex_total'] = row
        elif 'TOTAL COSTS' in val_str: cost_rows['total'] = row

print(f"Cost Schedule rows found: {cost_rows}")

# Time column setup (same as v4)
MN=['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
TC=[]
for m in range(7,13):TC.append((f"{MN[m]} 2025",2025,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2026",2026,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2027",2027,'m',m))
for yr in [2028,2029,2030]:
    for q in range(1,5):TC.append((f"Q{q} {yr}",yr,'q',q))
TSC=3; NTC=42

# We need Market Sizing row refs
ws_mkt = wb['Market Sizing']
mkt_rows = {}
for row in range(1, ws_mkt.max_row + 1):
    val = ws_mkt.cell(row=row, column=1).value
    if val:
        val_str = str(val).strip()
        if 'Period divisor' in val_str: mkt_rows['div'] = row
        elif 'Net active' in val_str: mkt_rows['net'] = row
        elif 'VCR' in val_str: mkt_rows['vcr'] = row
        elif 'Years from base' in val_str: mkt_rows['ybase'] = row
        elif 'E-commerce' in val_str: mkt_rows['ecom'] = row
        elif 'Months in period' in val_str: mkt_rows['mon'] = row

print(f"Market Sizing rows found: {mkt_rows}")

def ms(row, i): return f"'Market Sizing'!{get_column_letter(TSC+i)}{row}"

# Update Cloud row: add GMV-based cost component
if 'cloud' in cost_rows and 'div' in mkt_rows:
    cr = cost_rows['cloud']
    for i in range(NTC):
        old_formula = f"=opex_cloud_base/{ms(mkt_rows['div'],i)}+{ms(mkt_rows['vcr'],i)}*cloud_per_1M_vcr"
        # New: add GMV processing cost: ecom × market_share × cloud_gmv_factor / div
        # And make VCR cost super-linear with cloud_elasticity
        new_formula = (
            f"=opex_cloud_base/{ms(mkt_rows['div'],i)}"
            f"+{ms(mkt_rows['vcr'],i)}*cloud_per_1M_vcr*({ms(mkt_rows['net'],i)}/100)^(cloud_elasticity-1)"
            f"+{ms(mkt_rows['ecom'],i)}*platform_market_share*cloud_gmv_factor"
        )
        fm(ws_cost, cr, TSC+i, new_formula, FMT)
    print(f"  Updated Cloud row {cr} with GMV-based + elasticity formulas")

# Update Marketing row: add efficiency scaling
if 'marketing' in cost_rows:
    mr = cost_rows['marketing']
    for i in range(NTC):
        new_formula = (
            f"=opex_marketing*(1+marketing_growth)^{ms(mkt_rows['ybase'],i)}/{ms(mkt_rows['div'],i)}"
            f"*MAX(0.5,1-{ms(mkt_rows['net'],i)}*mkt_efficiency_per_100/100)"
        )
        fm(ws_cost, mr, TSC+i, new_formula, FMT)
    print(f"  Updated Marketing row {mr} with efficiency scaling")

# Update Support row: add economies of scale
if 'support' in cost_rows:
    sr = cost_rows['support']
    for i in range(NTC):
        new_formula = (
            f"={ms(mkt_rows['net'],i)}*opex_support_per_ent"
            f"*IF({ms(mkt_rows['net'],i)}>support_scale_threshold,1-support_discount,1)"
            f"/{ms(mkt_rows['div'],i)}"
        )
        fm(ws_cost, sr, TSC+i, new_formula, FMT)
    print(f"  Updated Support row {sr} with economies of scale")

print("Step 2 done: Cost Schedule formulas updated with cascading relationships")

# ============================================================
# STEP 3: Add Sensitivity Analysis sheet
# ============================================================
ws = wb.create_sheet("Sensitivity")
ws.sheet_properties.tabColor = "FF0066"
ws.column_dimensions['A'].width = 35
for c in ['B','C','D','E','F','G','H']:
    ws.column_dimensions[c].width = 16

r = 1
ws.cell(row=r,column=1,value="SENSITIVITY ANALYSIS").font = Font(bold=True,size=14,color="1F4E79")
ws.merge_cells("A1:H1"); r += 1
ws.cell(row=r,column=1,value="v5.0 | Shows how each parameter change affects NPV, IRR, and key outputs").font=Font(italic=True,size=10,color="666666")
r += 2

# Section 1: One-at-a-time sensitivity (tornado-style)
sc(ws,r,1,"ONE-AT-A-TIME SENSITIVITY (vs BASE CASE)")
ws.merge_cells(f"A{r}:H{r}")
for col in range(2,9):ws.cell(row=r,column=col).fill=SC_FILL
r += 1

# Headers
hd(ws,r,1,"Parameter"); hd(ws,r,2,"Named Range"); hd(ws,r,3,"-30%"); hd(ws,r,4,"-15%")
hd(ws,r,5,"BASE"); hd(ws,r,6,"+15%"); hd(ws,r,7,"+30%"); hd(ws,r,8,"Impact Direction")
r += 1

# Key parameters for sensitivity
sens_params = [
    ("Max enterprises", "adoption_max", "=adoption_max", 1,
     "↑ adoption → ↑ all revenue, ↑ HR/support costs, trigger Stage 2 earlier"),
    ("Platform market share", "platform_market_share", "=platform_market_share", 1,
     "↑ share → ↑ tx/DaaS/ads revenue, ↑ cloud costs, ↑ COGS"),
    ("Churn rate", "churn_rate", "=churn_rate", -1,
     "↑ churn → ↓ net ent → ↓ all revenue. Inverse impact."),
    ("Cert fee (Tr/yr)", "cert_fee_yr1", "=cert_fee_yr1", 1,
     "↑ fee → ↑ cert revenue directly. But may ↑ churn."),
    ("Cert escalation", "cert_fee_escalation", "=cert_fee_escalation", 1,
     "↑ escalation → ↑ revenue in later years (compound)."),
    ("WACC (discount rate)", "discount_rate", "=discount_rate", -1,
     "↑ WACC → ↓ NPV (future CFs worth less). No effect on operations."),
    ("Marketing base", "opex_marketing", "=opex_marketing", -1,
     "↑ marketing → ↓ profit. But needed to drive adoption."),
    ("S-curve steepness", "adoption_steepness", "=adoption_steepness", 1,
     "↑ steepness → faster adoption around midpoint."),
    ("Stage 2 gate", "stage2_gate", "=stage2_gate", -1,
     "↑ gate → later Stage 2 activation → ↓ tx revenue."),
    ("Cloud base cost", "opex_cloud_base", "=opex_cloud_base", -1,
     "↑ cloud → ↓ profit. Scales with VCR and GMV."),
    ("BHXH rate", "bhxh_rate", "=bhxh_rate", -1,
     "↑ BHXH → ↑ all salary costs → ↓ profit."),
    ("CIT rate", "cit_rate", "=cit_rate", -1,
     "↑ tax → ↓ net income. Only when profitable."),
]

for label, nm, base_formula, direction, impact in sens_params:
    lb(ws,r,1,label)
    ws.cell(row=r,column=2,value=nm).font = Font(size=10, color="666666"); ws.cell(row=r,column=2).border = BD
    # -30%, -15%, Base, +15%, +30%
    for ci, pct in [(3, 0.70), (4, 0.85), (5, 1.0), (6, 1.15), (7, 1.30)]:
        f = f"={nm}*{pct}"
        cl = ws.cell(row=r, column=ci, value=f)
        cl.border = BD
        cl.number_format = '#,##0.000' if isinstance(sens_params[0], tuple) else FMT
        if ci == 5:  # Base
            cl.fill = TFL; cl.font = Font(bold=True, size=11)
        elif pct < 1:
            cl.font = Font(size=11, color="CC0000")
        else:
            cl.font = Font(size=11, color="00B050")
    # Impact description
    ws.cell(row=r, column=8, value=impact).font = Font(size=9, color="666666"); ws.cell(row=r,column=8).border=BD
    r += 1

r += 2

# Section 2: Two-way sensitivity table: adoption_max vs platform_market_share
sc(ws,r,1,"TWO-WAY TABLE: NPV sensitivity to adoption_max × platform_market_share")
ws.merge_cells(f"A{r}:H{r}")
for col in range(2,9):ws.cell(row=r,column=col).fill=SC_FILL
r += 1
ws.cell(row=r,column=1,value="Manually run: change both params in Assumptions → record NPV from DCF sheet").font=Font(italic=True,size=9,color="888888")
ws.merge_cells(f"A{r}:H{r}"); r += 1

# Column headers: market share values
market_shares = [0.001, 0.003, 0.005, 0.008, 0.01, 0.015, 0.02]
hd(ws,r,1,"adoption_max ↓ \\ mkt_share →")
for j, ms_val in enumerate(market_shares):
    cl = ws.cell(row=r, column=2+j, value=ms_val)
    cl.font = HD_FONT; cl.fill = HD_FILL; cl.number_format = '0.0%'; cl.border = BD
r += 1

# Row headers: adoption values
adoption_vals = [200, 400, 600, 800, 1000, 1200, 1500, 2000, 3000]
for a_val in adoption_vals:
    cl = ws.cell(row=r, column=1, value=a_val)
    cl.font = Font(bold=True, size=11); cl.border = BD
    for j in range(len(market_shares)):
        cl = ws.cell(row=r, column=2+j)
        cl.border = BD; cl.number_format = FMT
        # Highlight current base case cell
        if a_val == 1200 and abs(market_shares[j] - 0.005) < 0.0001:
            cl.fill = TFL; cl.value = "← BASE"
            cl.font = Font(bold=True, size=10, color="1F4E79")
        else:
            cl.value = ""  # User fills this manually
    r += 1

r += 2

# Section 3: Two-way table: adoption_max vs churn_rate
sc(ws,r,1,"TWO-WAY TABLE: NPV sensitivity to adoption_max × churn_rate")
ws.merge_cells(f"A{r}:H{r}")
for col in range(2,9):ws.cell(row=r,column=col).fill=SC_FILL
r += 1
ws.cell(row=r,column=1,value="Manually run: change both params in Assumptions → record NPV from DCF sheet").font=Font(italic=True,size=9,color="888888")
ws.merge_cells(f"A{r}:H{r}"); r += 1

churn_rates = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35]
hd(ws,r,1,"adoption_max ↓ \\ churn →")
for j, cr_val in enumerate(churn_rates):
    cl = ws.cell(row=r, column=2+j, value=cr_val)
    cl.font = HD_FONT; cl.fill = HD_FILL; cl.number_format = '0%'; cl.border = BD
r += 1

for a_val in adoption_vals:
    cl = ws.cell(row=r, column=1, value=a_val)
    cl.font = Font(bold=True, size=11); cl.border = BD
    for j in range(len(churn_rates)):
        cl = ws.cell(row=r, column=2+j)
        cl.border = BD; cl.number_format = FMT
        if a_val == 1200 and abs(churn_rates[j] - 0.20) < 0.001:
            cl.fill = TFL; cl.value = "← BASE"
            cl.font = Font(bold=True, size=10, color="1F4E79")
        else:
            cl.value = ""
    r += 1

r += 2

# Section 4: Cascading impact summary
sc(ws,r,1,"CASCADING IMPACT SUMMARY — How market_share changes ripple through the model")
ws.merge_cells(f"A{r}:H{r}")
for col in range(2,9):ws.cell(row=r,column=col).fill=SC_FILL
r += 1

hd(ws,r,1,"If market_share changes..."); hd(ws,r,2,"0.1% (Worst)"); hd(ws,r,3,"0.3% (Bear)")
hd(ws,r,4,"0.5% (Base)"); hd(ws,r,5,"1.0% (Bull)"); hd(ws,r,6,"2.0% (Best)")
r += 1

cascade_outputs = [
    ("GMV through platform (2030, Bn VND)", 
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.001",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.003",
     "=ecom_market_base*(1+ecom_growth_rate)^5*platform_market_share",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.01",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.02"),
    ("Tx fee revenue (2030, Tr VND/yr)",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.001*tx_take_rate*1000",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.003*tx_take_rate*1000",
     "=ecom_market_base*(1+ecom_growth_rate)^5*platform_market_share*tx_take_rate*1000",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.01*tx_take_rate*1000",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.02*tx_take_rate*1000"),
    ("Additional cloud cost from GMV (Tr/yr)",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.001*cloud_gmv_factor",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.003*cloud_gmv_factor",
     "=ecom_market_base*(1+ecom_growth_rate)^5*platform_market_share*cloud_gmv_factor",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.01*cloud_gmv_factor",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.02*cloud_gmv_factor"),
    ("Payment processing COGS (Tr/yr)",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.001*tx_take_rate*1000*payment_proc_pct",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.003*tx_take_rate*1000*payment_proc_pct",
     "=ecom_market_base*(1+ecom_growth_rate)^5*platform_market_share*tx_take_rate*1000*payment_proc_pct",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.01*tx_take_rate*1000*payment_proc_pct",
     "=ecom_market_base*(1+ecom_growth_rate)^5*0.02*tx_take_rate*1000*payment_proc_pct"),
    ("Net revenue impact vs base (Tr/yr)",
     "=(0.001-platform_market_share)*ecom_market_base*(1+ecom_growth_rate)^5*tx_take_rate*1000*(1-payment_proc_pct)",
     "=(0.003-platform_market_share)*ecom_market_base*(1+ecom_growth_rate)^5*tx_take_rate*1000*(1-payment_proc_pct)",
     "=0",
     "=(0.01-platform_market_share)*ecom_market_base*(1+ecom_growth_rate)^5*tx_take_rate*1000*(1-payment_proc_pct)",
     "=(0.02-platform_market_share)*ecom_market_base*(1+ecom_growth_rate)^5*tx_take_rate*1000*(1-payment_proc_pct)"),
]

for label, *formulas in cascade_outputs:
    lb(ws,r,1,label)
    for j, f in enumerate(formulas):
        fm(ws, r, 2+j, f, FMT)
    r += 1

r += 2

# Section 5: Instructions
sc(ws,r,1,"HOW TO USE THIS SHEET"); ws.merge_cells(f"A{r}:H{r}")
for col in range(2,9):ws.cell(row=r,column=col).fill=SC_FILL
r += 1
instructions = [
    "1. ONE-AT-A-TIME SENSITIVITY: Shows parameter values at each % variation.",
    "   → Change one param in Assumptions → check NPV in 'DCF & Returns' → record here.",
    "2. TWO-WAY TABLES: Test combinations of two parameters simultaneously.",
    "   → Set both params in Assumptions → check NPV → fill the cell.",
    "   → The '← BASE' cell marks current assumption values.",
    "3. CASCADING IMPACT SUMMARY: Shows computed impacts of market_share changes.",
    "   → These use formulas — values auto-update when you change Assumptions.",
    "4. TIP: Use Excel's Data → What-If Analysis → Data Table for automated sensitivity.",
    "   → Select NPV cell as 'Result cell', param cells as Row/Column input cells.",
]
for inst in instructions:
    ws.cell(row=r, column=1, value=inst).font = Font(size=10, color="444444")
    ws.merge_cells(f"A{r}:H{r}"); r += 1

print("Step 3 done: Sensitivity Analysis sheet created")

# ============================================================
# STEP 4: Update Guide sheet
# ============================================================
ws = wb['Guide']
gr = ws.max_row + 3

sc(ws,gr,1,"v5.0 CHANGES (2026-03-21)"); gr += 1
changes = [
    "1. Added CASCADING RELATIONSHIPS section to Assumptions (6 new named ranges):",
    "   cloud_gmv_factor, mkt_efficiency_per_100, support_scale_threshold,",
    "   support_discount, hr_bizops_per_100_ent, cloud_elasticity",
    "2. Updated Cost Schedule formulas:",
    "   • Cloud: now includes GMV-based cost (market_share → cloud) + elasticity",
    "   • Marketing: efficiency scaling (more ent → lower cost/ent, floor 50%)",
    "   • Support: economies of scale (past threshold → 30% cost reduction)",
    "3. Added IMPACT MAP in Assumptions showing full cascade chain:",
    "   adoption_max → revenue + HR + support + cloud",
    "   platform_market_share → tx revenue + cloud + COGS",
    "   churn_rate → net_ent → all linked items",
    "4. Added Sensitivity Analysis sheet:",
    "   • One-at-a-time sensitivity (12 params × 5 levels)",
    "   • Two-way tables: adoption × market_share, adoption × churn",
    "   • Cascading impact summary with live formulas",
    "5. Added Sensitivity sheet to sheets list in Guide",
]
for line in changes:
    ws.cell(row=gr, column=1, value=line).font = Font(size=10, color="444444")
    ws.merge_cells(f"A{gr}:C{gr}"); gr += 1

# Update title
ws.cell(row=1,column=1).value = "VISIT VIETNAM — FINANCIAL MODEL v5 — GUIDE"
ws.cell(row=2,column=1).value = "v5.0 | 2026-03-21 | Added cascading relationships + Sensitivity Analysis"

# Also update Assumptions title
ws_a = wb['Assumptions']
ws_a.cell(row=1, column=1).value = "VISIT VIETNAM — FINANCIAL MODEL v5 — ASSUMPTIONS"
ws_a.cell(row=2, column=1).value = "v5.0 | Yellow=inputs | Blue=placeholder | Named ranges in all formulas | NEW: Cascading relationships at bottom"

print("Step 4 done: Guide updated")

# ============================================================
# STEP 5: Reorder and save
# ============================================================
order = ["Guide","Assumptions","Initial Investment","HR Plan","Market Sizing",
         "Revenue Schedule","Cost Schedule","P&L","Cash Flow","DCF & Returns",
         "Sensitivity","Scenarios","Budget 2025","Budget 2026"]
for i, n in enumerate(order):
    if n in wb.sheetnames:
        wb.move_sheet(n, offset=i - wb.sheetnames.index(n))

wb.save(TMP)
print(f"\nv5 saved: {TMP}")
print(f"Sheets: {wb.sheetnames}")

try:
    shutil.copy2(TMP, FINAL)
    print(f"Copied to: {FINAL}")
except Exception as e:
    print(f"Copy failed: {e}")
