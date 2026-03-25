"""
v6: Fix HR Plan with enterprise-driven BizOps + computed sensitivity NPV values
"""
import openpyxl, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

SRC = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\hhdlqg-financial-projection\VisitVietnam_FinancialModel_v5.xlsx"
TMP = r"E:\tmp\v6_build.xlsx"
FINAL = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\hhdlqg-financial-projection\VisitVietnam_FinancialModel_v6.xlsx"

shutil.copy2(SRC, TMP)
wb = openpyxl.load_workbook(TMP)

PH_FILL=PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid")
PH_FONT=Font(color="0000FF",size=11)
HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LBF=Font(size=11);TF=Font(bold=True,size=11)
TFL=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),
    top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FMT='#,##0';FP='0.0%'

def add_nm(name, sheet, row, col=2):
    ref = f"'{sheet}'!${get_column_letter(col)}${row}"
    try: wb.defined_names.delete(name)
    except: pass
    wb.defined_names.add(DefinedName(name, attr_text=ref))

def hd(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def sc(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
def lb(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LBF;cl.border=BD
def fm(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=Font(size=11);cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def tot(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=TF;cl.fill=TFL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)

MN=['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
TC=[]
for m in range(7,13):TC.append((f"{MN[m]} 2025",2025,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2026",2026,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2027",2027,'m',m))
for yr in [2028,2029,2030]:
    for q in range(1,5):TC.append((f"Q{q} {yr}",yr,'q',q))
TSC=3; NTC=42

# ============================================================
# STEP 1: Add bizops_floor named range to Assumptions
# ============================================================
ws_a = wb['Assumptions']
# Find last cascade param row
last_row = ws_a.max_row
# Find the cascade section - look for empty rows after cascade params
cascade_end = None
for row in range(1, last_row + 1):
    val = ws_a.cell(row=row, column=1).value
    if val and 'Cloud cost elasticity' in str(val):
        cascade_end = row
        break

if cascade_end:
    ins_r = cascade_end + 1
    lb(ws_a, ins_r, 1, "BizOps minimum floor headcount")
    cl = ws_a.cell(row=ins_r, column=2, value=2); cl.font=PH_FONT; cl.fill=PH_FILL; cl.border=BD; cl.number_format=FMT
    ws_a.cell(row=ins_r, column=3, value="people").font=LBF; ws_a.cell(row=ins_r,column=3).border=BD
    ws_a.cell(row=ins_r, column=4, value="MIN BizOps before scaling kicks in").font=Font(italic=True,size=10,color="888888"); ws_a.cell(row=ins_r,column=4).border=BD
    add_nm("bizops_floor", "Assumptions", ins_r)
    print(f"Added bizops_floor at row {ins_r}")

# Update title
ws_a.cell(row=1,column=1).value = "VISIT VIETNAM — FINANCIAL MODEL v6 — ASSUMPTIONS"

# ============================================================
# STEP 2: Update HR Plan BizOps headcount with formula
# ============================================================
ws_hr = wb['HR Plan']
# Find BizOps row and Market Sizing net_ent row
hr_rows = {}
for row in range(1, ws_hr.max_row + 1):
    val = ws_hr.cell(row=row, column=1).value
    if val:
        v = str(val).strip()
        if v == 'Biz Ops': hr_rows['bizops'] = row
        elif v == 'Management': hr_rows['mgmt'] = row
        elif v == 'Development': hr_rows['dev'] = row
        elif 'Data' in v and 'Analytics' in v: hr_rows['data'] = row
        elif 'TOTAL HEADCOUNT' in v: hr_rows['hc_total'] = row
        elif 'TOTAL SALARY' in v: hr_rows['sal_total'] = row
        elif 'Recruitment' in v: hr_rows['recruit'] = row
        elif 'TOTAL HR' in v: hr_rows['hr_total'] = row

print(f"HR Plan rows: {hr_rows}")

ws_mkt = wb['Market Sizing']
mkt_rows = {}
for row in range(1, ws_mkt.max_row + 1):
    val = ws_mkt.cell(row=row, column=1).value
    if val:
        v = str(val).strip()
        if 'Net active' in v: mkt_rows['net'] = row
        elif 'Period divisor' in v: mkt_rows['div'] = row
        elif 'Years from base' in v: mkt_rows['ybase'] = row
        elif 'VCR' in v: mkt_rows['vcr'] = row
        elif 'E-commerce' in v: mkt_rows['ecom'] = row
        elif 'Months in period' in v: mkt_rows['mon'] = row
        elif 'Year fraction' in v: mkt_rows['yfrac'] = row

print(f"Market Sizing rows: {mkt_rows}")

def ms(row, i): return f"'Market Sizing'!{get_column_letter(TSC+i)}{row}"

# Update BizOps row: MAX(bizops_floor, ROUND(net_ent × hr_bizops_per_100_ent / 100, 0))
if 'bizops' in hr_rows and 'net' in mkt_rows:
    br = hr_rows['bizops']
    for i in range(NTC):
        formula = f"=MAX(bizops_floor,ROUND({ms(mkt_rows['net'],i)}*hr_bizops_per_100_ent/100,0))"
        cl = ws_hr.cell(row=br, column=TSC+i, value=formula)
        cl.font = Font(size=11); cl.border = BD; cl.number_format = FMT
        # No longer yellow - it's a formula now, not a placeholder
    ws_hr.cell(row=br, column=1, value="Biz Ops (scaled)").font = LBF
    print(f"Updated BizOps row {br} with scaling formula")

# ============================================================
# STEP 3: Rebuild Sensitivity sheet with computed NPV
# ============================================================
# Delete old Sensitivity sheet
if 'Sensitivity' in wb.sheetnames:
    del wb['Sensitivity']

ws = wb.create_sheet("Sensitivity")
ws.sheet_properties.tabColor = "FF0066"
ws.column_dimensions['A'].width = 38
for c_l in 'BCDEFGHIJK':
    ws.column_dimensions[c_l].width = 14

r = 1
ws.cell(row=r,column=1,value="SENSITIVITY ANALYSIS — v6 (computed NPV)").font = Font(bold=True,size=14,color="1F4E79")
ws.merge_cells("A1:I1"); r += 2

# -----------------------------------------------------------
# Helper: Simplified NPV formula for a given (am, ms, cr) override
# This is an approximation using 6 annual periods
# -----------------------------------------------------------
# The NPV formula for each cell needs to compute:
# For years 0..5 (H2_2025 through 2030):
#   net_ent(y) = MAX(pilot, am / (1+EXP(-steepness*(base+0.5+y-midpoint)))) * (1-cr)
#   rev(y) = net * (cert_fee*(1+esc)^y + training_pct*training_fee + audit_fee)
#           + IF(net >= gate, ecom*(1+g)^y * ms * take * 1000, 0)
#   cost(y) ≈ capex_total*maint% + opex_total_approx
#   fcf(y) = rev(y) - cost(y) - IF(y==0, capex_total, 0)
# NPV = SUM( fcf(y) / (1+wacc)^y )

# To make this work in Excel formulas, I'll build helper columns
# that compute net_ent and revenue for each test case scenario.

# SECTION 1: One-at-a-time tornado with COMPUTED impact
sc(ws,r,1,"ONE-AT-A-TIME SENSITIVITY — Approximate NPV Impact (Tr VND)")
ws.merge_cells(f"A{r}:I{r}")
for col in range(2,10):ws.cell(row=r,column=col).fill=SC_FILL
r += 1

hd(ws,r,1,"Parameter"); hd(ws,r,2,"-30%"); hd(ws,r,3,"-15%"); hd(ws,r,4,"BASE")
hd(ws,r,5,"+15%"); hd(ws,r,6,"+30%"); hd(ws,r,7,"Value @Base"); hd(ws,r,8,"Impact ↑ or ↓")
r += 1

# For each parameter, I compute a simplified annual-average approach:
# avg_net = adoption_max / (1+EXP(-steepness*(2028-midpoint))) * (1-churn)  ← approx midpoint
# total_5yr_revenue ≈ avg_net × (cert + training_pct*training + audit) × 5
#                    + IF(avg_net>=gate, ecom_avg × ms × take × 1000 × 3, 0)  ← ~3 years of stage 2
# total_5yr_cost ≈ (marketing + cloud + hr + other) × 5 + capex
# simplified_NPV ≈ (revenue - cost) / (1 + wacc)^2.5  ← midpoint discount

# I'll define a mega-formula helper function
def npv_formula(am_override=None, ms_override=None, cr_override=None,
                cf_override=None, ce_override=None, dr_override=None,
                mk_override=None, st_override=None, sg_override=None,
                cb_override=None, bh_override=None, ct_override=None):
    """Build simplified NPV formula with parameter overrides."""
    am = am_override or "adoption_max"
    msh = ms_override or "platform_market_share"
    cr = cr_override or "churn_rate"
    cf = cf_override or "cert_fee_yr1"
    ce = ce_override or "cert_fee_escalation"
    dr = dr_override or "discount_rate"
    mk = mk_override or "opex_marketing"
    st = st_override or "adoption_steepness"
    sg = sg_override or "stage2_gate"
    cb = cb_override or "opex_cloud_base"
    bh = bh_override or "bhxh_rate"
    ct = ct_override or "cit_rate"

    # avg_net at midpoint ≈ am / (1+EXP(-st*(2028-adoption_midpoint))) * (1-cr)
    avg_net = f"({am}/(1+EXP(-{st}*(2028-adoption_midpoint)))*(1-{cr}))"

    # 5-year total Stage 1 revenue (cert escalates, training/audit flat)
    # avg cert fee ≈ cert_fee × (1+esc)^2.5 (midpoint of 5 years)
    s1_rev = f"({avg_net}*({cf}*(1+{ce})^2.5+training_pct*training_fee+audit_fee)*5)"

    # Stage 2 revenue: ~3 years active (2028-2030), ecom at year 4 average
    # ecom_avg ≈ ecom_base * (1+growth)^4
    s2_rev = f"(IF({avg_net}>={sg},ecom_market_base*(1+ecom_growth_rate)^4*{msh}*tx_take_rate*1000*3,0))"

    # COGS
    cogs = f"({s2_rev}*payment_proc_pct)"

    # Total costs over 5 years
    # HR ≈ (16+22+30+41+55)/5 people avg × 30M avg salary × 12 × (1+bhxh) × (1+esc)^2 ≈
    hr_cost = f"(33*30*12*(1+{bh})*(1+salary_escalation)^2*5)"
    mkt_cost = f"({mk}*(1+marketing_growth)^2*5)"
    cloud_cost = f"({cb}*5+{avg_net}*vcr_per_enterprise*12*cloud_per_1M_vcr/1000000*5)"
    other_cost = f"((opex_branding+opex_external+opex_legal+opex_security+opex_partner_mgmt)*5)"
    maintenance = f"(capex_total*capex_maintenance_pct*4)"  # 4 years of maintenance

    total_cost = f"(capex_total+{hr_cost}+{mkt_cost}+{cloud_cost}+{other_cost}+{maintenance})"

    # Pre-tax profit
    pretax = f"({s1_rev}+{s2_rev}-{cogs}-{total_cost})"

    # After-tax
    after_tax = f"({pretax}*(1-{ct}))"

    # Simplified NPV ≈ after_tax / (1+wacc)^2.5
    npv = f"={after_tax}/(1+{dr})^2.5"

    return npv

# One-at-a-time rows
oat_params = [
    ("Max enterprises (adoption_max)", "adoption_max",
     lambda p: npv_formula(am_override=f"adoption_max*{p}"),
     "=adoption_max", "↑ adoption →  ↑ all revenue"),
    ("Market share (platform_market_share)", "platform_market_share",
     lambda p: npv_formula(ms_override=f"platform_market_share*{p}"),
     "=platform_market_share", "↑ share → ↑ Stage 2 rev"),
    ("Churn rate", "churn_rate",
     lambda p: npv_formula(cr_override=f"churn_rate*{p}"),
     "=churn_rate", "↑ churn → ↓ net ent"),
    ("Cert fee (Tr/yr)", "cert_fee_yr1",
     lambda p: npv_formula(cf_override=f"cert_fee_yr1*{p}"),
     "=cert_fee_yr1", "↑ fee → ↑ Stage 1 rev"),
    ("Cert escalation", "cert_fee_escalation",
     lambda p: npv_formula(ce_override=f"cert_fee_escalation*{p}"),
     "=cert_fee_escalation", "↑ esc → ↑ later rev"),
    ("WACC (discount_rate)", "discount_rate",
     lambda p: npv_formula(dr_override=f"discount_rate*{p}"),
     "=discount_rate", "↑ WACC → ↓ NPV"),
    ("Marketing base", "opex_marketing",
     lambda p: npv_formula(mk_override=f"opex_marketing*{p}"),
     "=opex_marketing", "↑ mkt → ↓ profit"),
    ("S-curve steepness", "adoption_steepness",
     lambda p: npv_formula(st_override=f"adoption_steepness*{p}"),
     "=adoption_steepness", "↑ steep → faster adopt"),
    ("Stage 2 gate", "stage2_gate",
     lambda p: npv_formula(sg_override=f"stage2_gate*{p}"),
     "=stage2_gate", "↑ gate → later Stage 2"),
    ("Cloud base cost", "opex_cloud_base",
     lambda p: npv_formula(cb_override=f"opex_cloud_base*{p}"),
     "=opex_cloud_base", "↑ cloud → ↓ profit"),
    ("BHXH rate", "bhxh_rate",
     lambda p: npv_formula(bh_override=f"bhxh_rate*{p}"),
     "=bhxh_rate", "↑ BHXH → ↑ HR cost"),
    ("CIT rate", "cit_rate",
     lambda p: npv_formula(ct_override=f"cit_rate*{p}"),
     "=cit_rate", "↑ tax → ↓ net income"),
]

for label, nm, fn, base_ref, impact in oat_params:
    lb(ws,r,1,label)
    for ci, pct in [(2, 0.70), (3, 0.85), (4, 1.0), (5, 1.15), (6, 1.30)]:
        formula = fn(pct)
        cl = ws.cell(row=r, column=ci, value=formula)
        cl.border = BD; cl.number_format = FMT
        if ci == 4:
            cl.fill = TFL; cl.font = Font(bold=True, size=11)
        elif pct < 1:
            cl.font = Font(size=11, color="CC0000")
        else:
            cl.font = Font(size=11, color="00B050")
    fm(ws,r,7, base_ref, '#,##0.000')
    ws.cell(row=r, column=8, value=impact).font=Font(size=9,color="666666"); ws.cell(row=r,column=8).border=BD
    r += 1

r += 2

# -----------------------------------------------------------
# SECTION 2: Two-way table adoption × market_share with COMPUTED NPV
# -----------------------------------------------------------
sc(ws,r,1,"TWO-WAY NPV: adoption_max × platform_market_share (Tr VND)")
ws.merge_cells(f"A{r}:I{r}")
for col in range(2,10):ws.cell(row=r,column=col).fill=SC_FILL
r += 1

market_shares = [0.001, 0.003, 0.005, 0.008, 0.01, 0.015, 0.02]
hd(ws,r,1,"adoption ↓ \\ mkt_share →")
for j, ms_val in enumerate(market_shares):
    cl = ws.cell(row=r, column=2+j, value=ms_val)
    cl.font=HD_FONT; cl.fill=HD_FILL; cl.number_format='0.0%'; cl.border=BD
r += 1

adoption_vals = [200, 400, 600, 800, 1000, 1200, 1500, 2000, 3000]
for a_val in adoption_vals:
    cl = ws.cell(row=r, column=1, value=a_val)
    cl.font = Font(bold=True, size=11); cl.border = BD
    for j, ms_val in enumerate(market_shares):
        formula = npv_formula(am_override=str(a_val), ms_override=str(ms_val))
        cl = ws.cell(row=r, column=2+j, value=formula)
        cl.border = BD; cl.number_format = FMT
        if a_val == 1200 and abs(ms_val - 0.005) < 0.0001:
            cl.fill = TFL; cl.font = Font(bold=True, size=11)
    r += 1

r += 2

# -----------------------------------------------------------
# SECTION 3: Two-way table adoption × churn with COMPUTED NPV
# -----------------------------------------------------------
sc(ws,r,1,"TWO-WAY NPV: adoption_max × churn_rate (Tr VND)")
ws.merge_cells(f"A{r}:I{r}")
for col in range(2,10):ws.cell(row=r,column=col).fill=SC_FILL
r += 1

churn_rates = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35]
hd(ws,r,1,"adoption ↓ \\ churn →")
for j, cr_val in enumerate(churn_rates):
    cl = ws.cell(row=r, column=2+j, value=cr_val)
    cl.font=HD_FONT; cl.fill=HD_FILL; cl.number_format='0%'; cl.border=BD
r += 1

for a_val in adoption_vals:
    cl = ws.cell(row=r, column=1, value=a_val)
    cl.font = Font(bold=True, size=11); cl.border = BD
    for j, cr_val in enumerate(churn_rates):
        formula = npv_formula(am_override=str(a_val), cr_override=str(cr_val))
        cl = ws.cell(row=r, column=2+j, value=formula)
        cl.border = BD; cl.number_format = FMT
        if a_val == 1200 and abs(cr_val - 0.20) < 0.001:
            cl.fill = TFL; cl.font = Font(bold=True, size=11)
    r += 1

r += 2

# -----------------------------------------------------------
# SECTION 4: Two-way table market_share × discount_rate
# -----------------------------------------------------------
sc(ws,r,1,"TWO-WAY NPV: platform_market_share × discount_rate (Tr VND)")
ws.merge_cells(f"A{r}:I{r}")
for col in range(2,10):ws.cell(row=r,column=col).fill=SC_FILL
r += 1

wacc_rates = [0.08, 0.10, 0.12, 0.15, 0.18, 0.20, 0.25]
hd(ws,r,1,"mkt_share ↓ \\ WACC →")
for j, w_val in enumerate(wacc_rates):
    cl = ws.cell(row=r, column=2+j, value=w_val)
    cl.font=HD_FONT; cl.fill=HD_FILL; cl.number_format='0%'; cl.border=BD
r += 1

ms_test_vals = [0.001, 0.003, 0.005, 0.008, 0.01, 0.015, 0.02]
for ms_val in ms_test_vals:
    cl = ws.cell(row=r, column=1, value=ms_val)
    cl.font = Font(bold=True, size=11); cl.border = BD; cl.number_format = '0.0%'
    for j, w_val in enumerate(wacc_rates):
        formula = npv_formula(ms_override=str(ms_val), dr_override=str(w_val))
        cl = ws.cell(row=r, column=2+j, value=formula)
        cl.border = BD; cl.number_format = FMT
        if abs(ms_val - 0.005) < 0.0001 and abs(w_val - 0.15) < 0.001:
            cl.fill = TFL; cl.font = Font(bold=True, size=11)
    r += 1

r += 2

# Notes
sc(ws,r,1,"NOTES"); ws.merge_cells(f"A{r}:I{r}")
for col in range(2,10):ws.cell(row=r,column=col).fill=SC_FILL
r += 1
notes = [
    "• All NPV values are APPROXIMATE (simplified 5-year model using midpoint averages).",
    "• For exact NPV: change values in Assumptions sheet → read 'DCF & Returns' NPV cell.",
    "• Green highlighted cells = BASE CASE (current assumptions).",
    "• All formulas use named ranges from Assumptions — change a param → tables recalculate.",
    "• HR avg headcount estimated at 33 (weighted average 2025-2030). Actual model uses period-by-period.",
    "• Stage 2 revenue assumed active for ~3 years (2028-2030) when enterprises exceed gate.",
]
for n in notes:
    ws.cell(row=r,column=1,value=n).font=Font(size=9,color="666666")
    ws.merge_cells(f"A{r}:I{r}"); r += 1

# ============================================================
# STEP 4: Update Guide
# ============================================================
ws_g = wb['Guide']
ws_g.cell(row=1,column=1).value = "VISIT VIETNAM — FINANCIAL MODEL v6 — GUIDE"
ws_g.cell(row=2,column=1).value = "v6.0 | 2026-03-21 | HR scaling + computed sensitivity NPV"

gr = ws_g.max_row + 2
sc(ws_g,gr,1,"v6.0 CHANGES"); gr += 1
for line in [
    "1. HR Plan: BizOps headcount now formula-driven:",
    "   =MAX(bizops_floor, ROUND(net_enterprises × hr_bizops_per_100_ent/100, 0))",
    "   More enterprises → more BizOps staff automatically",
    "2. Added bizops_floor named range (=2) as minimum headcount",
    "3. Sensitivity sheet fully rebuilt with COMPUTED NPV formulas:",
    "   • One-at-a-time: 12 params × 5 levels, each cell = simplified NPV formula",
    "   • Two-way table 1: adoption_max(9) × market_share(7) = 63 NPV values",
    "   • Two-way table 2: adoption_max(9) × churn_rate(7) = 63 NPV values",
    "   • Two-way table 3: market_share(7) × WACC(7) = 49 NPV values",
    "   • All cells contain formulas with named ranges — auto-update when Assumptions change",
]:
    ws_g.cell(row=gr,column=1,value=line).font=Font(size=10,color="444444")
    ws_g.merge_cells(f"A{gr}:C{gr}"); gr += 1

# Reorder
order = ["Guide","Assumptions","Initial Investment","HR Plan","Market Sizing",
         "Revenue Schedule","Cost Schedule","P&L","Cash Flow","DCF & Returns",
         "Sensitivity","Scenarios","Budget 2025","Budget 2026"]
for i, n in enumerate(order):
    if n in wb.sheetnames:
        wb.move_sheet(n, offset=i - wb.sheetnames.index(n))

wb.save(TMP)
print(f"v6 saved: {TMP}")
print(f"Sheets: {wb.sheetnames}")

try:
    shutil.copy2(TMP, FINAL)
    print(f"Copied to: {FINAL}")
except Exception as e:
    print(f"Copy failed: {e}")
