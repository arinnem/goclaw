"""
v4 Part 2: Initial Investment + HR Plan + Market Sizing
"""
import openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

with open('/tmp/v4_state.json') as f: state=json.load(f)
OUT=state['out']; named=state['named']

PH_FILL=PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid")
PH_FONT=Font(color="0000FF",size=11)
HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LBF=Font(size=11);TF=Font(bold=True,size=11)
TFL=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FMT='#,##0'; FP='0.0%'; FD='0.000'

def hd(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=HD_FONT;cl.fill=HD_FILL;cl.alignment=Alignment(horizontal='center');cl.border=BD
def sc(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=SC_FONT;cl.fill=SC_FILL;cl.border=BD
def lb(ws,r,c,v):cl=ws.cell(row=r,column=c,value=v);cl.font=LBF;cl.border=BD
def fm(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=Font(size=11);cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def tot(ws,r,c,f,fmt=None):cl=ws.cell(row=r,column=c,value=f);cl.font=TF;cl.fill=TFL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)
def ph(ws,r,c,v,fmt=None):cl=ws.cell(row=r,column=c,value=v);cl.font=PH_FONT;cl.fill=PH_FILL;cl.border=BD;(setattr(cl,'number_format',fmt) if fmt else None)

MN=['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
TC=[]
for m in range(7,13):TC.append((f"{MN[m]} 2025",2025,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2026",2026,'m',m))
for m in range(1,13):TC.append((f"{MN[m]} 2027",2027,'m',m))
for yr in [2028,2029,2030]:
    for q in range(1,5):TC.append((f"Q{q} {yr}",yr,'q',q))
TSC=3;NTC=len(TC)

wb=openpyxl.load_workbook(OUT)

def add_nm(name,sheet,row,col=2):
    ref=f"'{sheet}'!${get_column_letter(col)}${row}"
    dn=DefinedName(name,attr_text=ref);wb.defined_names.add(dn)
    named[name]=f"{sheet}:{row}"

# ===== INITIAL INVESTMENT =====
ws=wb.create_sheet("Initial Investment");ws.sheet_properties.tabColor="C00000"
ws.column_dimensions['A'].width=40;ws.column_dimensions['B'].width=18;ws.column_dimensions['C'].width=12;ws.column_dimensions['D'].width=40
r=1;sc(ws,r,1,"INITIAL INVESTMENT — CAPEX BREAKDOWN (M VND)");r+=2
hd(ws,r,1,"CAPEX Item");hd(ws,r,2,"Amount (M VND)");hd(ws,r,3,"Timing");hd(ws,r,4,"Notes");r+=1

items=[
    ("capex_cloud_setup","Cloud infrastructure setup",800,"H2 2025","AWS/GCP initial"),
    ("capex_backend","Backend development",1200,"H2 2025","API, DB, core"),
    ("capex_frontend","Frontend + mobile",900,"H2 2025","Web + mobile"),
    ("capex_data_platform","Data platform",600,"H2 2025","Pipeline, analytics"),
    ("capex_security_setup","Security infrastructure",250,"H2 2025","Pentest, auth"),
    ("capex_ux_design","UX/UI design",350,"H2 2025","Design system"),
    ("capex_integration","3rd-party integrations",400,"H2 2025","Payment, OTA APIs"),
    ("capex_devops","DevOps & CI/CD",280,"H2 2025","Deploy pipeline"),
    ("capex_testing","QA & testing",200,"H2 2025","Auto + load test"),
    ("capex_contingency","CAPEX contingency (10%)",300,"H2 2025","Buffer"),
]
first=r
for nm,lbl,val,timing,note in items:
    lb(ws,r,1,lbl);ph(ws,r,2,val,FMT)
    ws.cell(row=r,column=3,value=timing).font=LBF;ws.cell(row=r,column=3).border=BD
    ws.cell(row=r,column=4,value=note).font=Font(italic=True,size=10,color="888888");ws.cell(row=r,column=4).border=BD
    add_nm(nm,"Initial Investment",r)
    r+=1
last=r-1;r+=1
lb(ws,r,1,"TOTAL INITIAL CAPEX")
tot(ws,r,2,f"=SUM(B{first}:B{last})",FMT)
add_nm("capex_total","Initial Investment",r)
cap_tot_r=r;r+=2

# Depreciation
sc(ws,r,1,"DEPRECIATION SCHEDULE");r+=1
hd(ws,r,1,"Year");hd(ws,r,2,"Annual D&A");hd(ws,r,3,"Accum D&A");hd(ws,r,4,"Net Book Value");r+=1
for i in range(6):
    yr=2025+i;lb(ws,r,1,str(yr))
    if i<5:fm(ws,r,2,"=capex_total/depreciation_years",FMT)
    else:fm(ws,r,2,"=0",FMT)
    if i==0:fm(ws,r,3,f"=B{r}",FMT)
    else:fm(ws,r,3,f"=C{r-1}+B{r}",FMT)
    fm(ws,r,4,f"=capex_total-C{r}",FMT)
    r+=1

# ===== HR PLAN =====
ws=wb.create_sheet("HR Plan");ws.sheet_properties.tabColor="7030A0"
ws.column_dimensions['A'].width=30;ws.column_dimensions['B'].width=14
r=1;sc(ws,r,1,"HR PLAN — HEADCOUNT & COST");r+=2

# Time headers
hd(ws,r,1,"Dept / Item");hd(ws,r,2,"Unit")
for i,(lbl,yr,tp,mp) in enumerate(TC):
    col=TSC+i;hd(ws,r,col,lbl);ws.column_dimensions[get_column_letter(col)].width=11
r+=1

sc(ws,r,1,"HEADCOUNT")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1

depts=[
    ("Management","salary_mgmt",[2]*6+[2]*12+[3]*12+[3]*4+[4]*4+[4]*4),
    ("Development","salary_dev",[12]*6+[12]*12+[14]*12+[16]*4+[18]*4+[20]*4),
    ("Biz Ops","salary_bizops",[2]*6+[6]*12+[10]*12+[18]*4+[28]*4+[35]*4),
    ("Data & Analytics","salary_data",[0]*6+[2]*12+[3]*12+[4]*4+[5]*4+[6]*4),
]
dept_rows={}
for dn,sal_nm,counts in depts:
    lb(ws,r,1,dn);ws.cell(row=r,column=2,value="people").font=LBF
    for i in range(NTC):ph(ws,r,TSC+i,counts[i],'#,##0')
    dept_rows[dn]=r;r+=1

hc_tot_r=r
lb(ws,r,1,"TOTAL HEADCOUNT");ws.cell(row=r,column=2,value="people").font=TF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    tot(ws,r,TSC+i,f"={cl}{dept_rows['Management']}+{cl}{dept_rows['Development']}+{cl}{dept_rows['Biz Ops']}+{cl}{dept_rows['Data & Analytics']}",FMT)
r+=2

# New hires
sc(ws,r,1,"NEW HIRES")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1
new_h_r=r
lb(ws,r,1,"New hires");ws.cell(row=r,column=2,value="people").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    if i==0:fm(ws,r,TSC+i,f"={cl}{hc_tot_r}",FMT)
    else:
        prev=get_column_letter(TSC+i-1)
        fm(ws,r,TSC+i,f"=MAX(0,{cl}{hc_tot_r}-{prev}{hc_tot_r})",FMT)
r+=2

# Salary costs
sc(ws,r,1,"SALARY COST (M VND)")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1

sal_rows={}
for dn,sal_nm,_ in depts:
    lb(ws,r,1,f"{dn} cost");ws.cell(row=r,column=2,value="M VND").font=LBF
    for i,(lbl,yr,tp,mp) in enumerate(TC):
        col=TSC+i;cl=get_column_letter(col)
        ye=yr-2025
        mult=1 if tp=='m' else 3
        # headcount * named_salary * months * (1+escalation)^years * (1+bhxh)
        fm(ws,r,col,f"={cl}{dept_rows[dn]}*{sal_nm}*{mult}*(1+salary_escalation)^{ye}*(1+bhxh_rate)",FMT)
    sal_rows[dn]=r;r+=1

sal_tot_r=r
lb(ws,r,1,"TOTAL SALARY");ws.cell(row=r,column=2,value="M VND").font=TF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    tot(ws,r,TSC+i,f"={cl}{sal_rows['Management']}+{cl}{sal_rows['Development']}+{cl}{sal_rows['Biz Ops']}+{cl}{sal_rows['Data & Analytics']}",FMT)
r+=1

rec_r=r
lb(ws,r,1,"Recruitment");ws.cell(row=r,column=2,value="M VND").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    fm(ws,r,TSC+i,f"={cl}{new_h_r}*30*12*recruit_cost_pct",FMT)
r+=1

hr_tot_r=r
lb(ws,r,1,"TOTAL HR COST");ws.cell(row=r,column=2,value="M VND").font=TF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    tot(ws,r,TSC+i,f"={cl}{sal_tot_r}+{cl}{rec_r}",FMT)

# ===== MARKET SIZING =====
ws=wb.create_sheet("Market Sizing");ws.sheet_properties.tabColor="4472C4"
ws.column_dimensions['A'].width=35;ws.column_dimensions['B'].width=10
r=1;sc(ws,r,1,"MARKET SIZING & ADOPTION");r+=2

# Period index
idx_r=r
lb(ws,r,1,"Period index");ws.cell(row=r,column=2,value="idx").font=LBF
for i in range(NTC):fm(ws,r,TSC+i,i,FMT)
r+=1

# Year fraction — calculated from base_year, start_month
yfrac_r=r
lb(ws,r,1,"Year fraction");ws.cell(row=r,column=2,value="year").font=LBF
for i,(lbl,yr,tp,mp) in enumerate(TC):
    col=TSC+i;cl=get_column_letter(col)
    if tp=='m':
        fm(ws,r,col,f"=base_year+(start_month-1+{cl}{idx_r})/12",FD)
    else:
        q_month=(i-30)*3+1
        fm(ws,r,col,f"=2028+({q_month}-0.5)/12",FD)
r+=1

ybase_r=r
lb(ws,r,1,"Years from base");ws.cell(row=r,column=2,value="yrs").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    fm(ws,r,TSC+i,f"=FLOOR({cl}{yfrac_r}-base_year,1)",FMT)
r+=1

div_r=r
lb(ws,r,1,"Period divisor");ws.cell(row=r,column=2,value="div").font=LBF
for i,(lbl,yr,tp,mp) in enumerate(TC):
    fm(ws,r,TSC+i,12 if tp=='m' else 4,FMT)
r+=1

mon_r=r
lb(ws,r,1,"Months in period");ws.cell(row=r,column=2,value="mon").font=LBF
for i,(lbl,yr,tp,mp) in enumerate(TC):
    fm(ws,r,TSC+i,1 if tp=='m' else 3,FMT)
r+=2

# Time headers
hd(ws,r,1,"Metric");hd(ws,r,2,"Unit")
for i,(lbl,yr,tp,mp) in enumerate(TC):
    col=TSC+i;hd(ws,r,col,lbl);ws.column_dimensions[get_column_letter(col)].width=11
r+=1

# E-com market
ecom_r=r
lb(ws,r,1,"E-commerce tourism market");ws.cell(row=r,column=2,value="Bn VND").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    fm(ws,r,TSC+i,f"=ecom_market_base*(1+ecom_growth_rate)^{cl}{ybase_r}/{cl}{div_r}",'#,##0.0')
r+=2

# S-Curve
sc(ws,r,1,"S-CURVE ADOPTION")
for i in range(NTC):ws.cell(row=r,column=TSC+i).fill=SC_FILL
r+=1

gross_r=r
lb(ws,r,1,"Gross enterprises");ws.cell(row=r,column=2,value="cnt").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    fm(ws,r,TSC+i,f"=MAX(pilot_count,adoption_max/(1+EXP(-adoption_steepness*({cl}{yfrac_r}-adoption_midpoint))))",FMT)
r+=1

churn_r=r
lb(ws,r,1,"Churned");ws.cell(row=r,column=2,value="cnt").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    if i==0:fm(ws,r,TSC+i,"=0",FMT)
    else:fm(ws,r,TSC+i,f"=ROUND({cl}{gross_r}*churn_rate/{cl}{div_r},0)",FMT)
r+=1

net_r=r
lb(ws,r,1,"Net active enterprises");ws.cell(row=r,column=2,value="cnt").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    tot(ws,r,TSC+i,f"={cl}{gross_r}-{cl}{churn_r}",FMT)
r+=1

new_ent_r=r
lb(ws,r,1,"New enterprises");ws.cell(row=r,column=2,value="cnt").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    if i==0:fm(ws,r,TSC+i,f"={cl}{net_r}",FMT)
    else:fm(ws,r,TSC+i,f"=MAX(0,{cl}{net_r}-{get_column_letter(TSC+i-1)}{net_r})",FMT)
r+=1

vcr_r=r
lb(ws,r,1,"VCR (M records)");ws.cell(row=r,column=2,value="M").font=LBF
for i in range(NTC):
    cl=get_column_letter(TSC+i)
    fm(ws,r,TSC+i,f"={cl}{net_r}*vcr_per_enterprise*{cl}{mon_r}/1000000",'0.00')

# Save state
state['hr_tot_r']=hr_tot_r;state['new_h_r']=new_h_r;state['hc_tot_r']=hc_tot_r
state['ecom_r']=ecom_r;state['net_r']=net_r;state['vcr_r']=vcr_r
state['yfrac_r']=yfrac_r;state['ybase_r']=ybase_r;state['div_r']=div_r;state['mon_r']=mon_r
state['named']=named;state['dept_rows']={k:v for k,v in dept_rows.items()}
with open('/tmp/v4_state.json','w') as f:json.dump(state,f)
wb.save(OUT)
print(f"Part 2 done: Initial Investment + HR Plan + Market Sizing")
