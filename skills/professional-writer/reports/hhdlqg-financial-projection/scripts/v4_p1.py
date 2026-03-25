"""
v4 Part 1: Complete Assumptions + Named Ranges (all 50 params, no row insertions)
"""
import openpyxl, json, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

PH_FILL=PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid")
PH_FONT=Font(color="0000FF",size=11)
HD_FILL=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
HD_FONT=Font(color="FFFFFF",bold=True,size=11)
SC_FILL=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
SC_FONT=Font(bold=True,size=11,color="1F4E79")
LB=Font(size=11); BD=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FMT_MN='#,##0';FMT_PCT='0.0%';FMT_INT='#,##0';FMT_DEC='0.0'
NOTE=Font(italic=True,size=10,color="888888")

OUT=r"E:\tmp\v4_build.xlsx"
FINAL=r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\hhdlqg-financial-projection\VisitVietnam_FinancialModel_v4.xlsx"

wb=openpyxl.Workbook()
named={}

def add_nm(name,row):
    ref=f"Assumptions!$B${row}"
    dn=DefinedName(name,attr_text=ref)
    wb.defined_names.add(dn)
    named[name]=row

ws=wb.active;ws.title="Assumptions";ws.sheet_properties.tabColor="FFD700"
ws.column_dimensions['A'].width=42;ws.column_dimensions['B'].width=18
ws.column_dimensions['C'].width=12;ws.column_dimensions['D'].width=45

r=1
c=ws.cell(row=r,column=1,value="VISIT VIETNAM — FINANCIAL MODEL v4 — ASSUMPTIONS");c.font=Font(bold=True,size=14,color="1F4E79")
ws.merge_cells("A1:D1");r+=1
ws.cell(row=r,column=1,value="v4.0 | All yellow=editable inputs | Blue font=placeholder | Named ranges in all formulas").font=Font(italic=True,size=10,color="666666");r+=2

# ALL params in one go — complete list, correct order, no insertions later
sections=[
    ("TIMING",[
        ("base_year","Base year",2025,"Year","Project start year",FMT_INT),
        ("start_month","Start month",7,"Month","July=7",FMT_INT),
    ]),
    ("MARKET DATA",[
        ("TAM_enterprises","Total tourism enterprises",40000,"","GSO/VNAT",FMT_INT),
        ("SAM_enterprises","Serviceable enterprises",15000,"","Digitally active",FMT_INT),
        ("ecom_market_base","E-commerce tourism market (base yr)",82500,"Bn VND","e-Conomy SEA",FMT_MN),
        ("ecom_growth_rate","E-tourism YoY growth",0.08,"%","Historical",FMT_PCT),
    ]),
    ("ADOPTION S-CURVE",[
        ("adoption_max","Maximum enterprises (ceiling)",1200,"","S-curve asymptote (v3: was 3000)",FMT_INT),
        ("adoption_midpoint","Midpoint year",2028,"Year","Inflection point",FMT_INT),
        ("adoption_steepness","Steepness factor",0.9,"","v3: was 1.2",FMT_DEC),
        ("pilot_count","Starting enterprises",15,"","Pilot partners",FMT_INT),
        ("churn_rate","Annual non-renewal rate",0.20,"%","v3: was 0.15",FMT_PCT),
    ]),
    ("PRICING",[
        ("cert_fee_yr1","Certification fee/enterprise/yr",25,"M VND","v3: was 30M",FMT_INT),
        ("cert_fee_escalation","Cert fee annual increase",0.05,"%","v3: was 0.08",FMT_PCT),
        ("tx_take_rate","Transaction fee (% of GMV)",0.02,"%","Platform commission",FMT_PCT),
        ("platform_market_share","% e-tourism GMV via VV",0.005,"%","v3: was 0.02",FMT_PCT),
        ("daas_price","DaaS price per M VCR",30,"M VND","v3: was 50",FMT_INT),
        ("ad_cpm","Advertising CPM",150,"K VND","B2B benchmark",FMT_INT),
        ("ad_impressions_base","Monthly impressions (K)",500,"K","Scales w/ enterprises",FMT_INT),
        ("training_fee","Training fee per engagement",15,"M VND","NEW stream",FMT_INT),
        ("audit_fee","Audit fee per certification",10,"M VND","NEW stream",FMT_INT),
        ("training_pct","% enterprises buying training",0.30,"%","Estimated",FMT_PCT),
    ]),
    ("COST PARAMETERS",[
        ("capex_maintenance_pct","Maintenance (% of CAPEX/yr)",0.18,"%","Industry 15-20%",FMT_PCT),
        ("opex_marketing","Marketing (annual base)",7000,"M VND","v3: was 6000",FMT_MN),
        ("marketing_growth","Marketing annual growth",0.15,"%","v3: was 0.10",FMT_PCT),
        ("opex_cloud_base","Cloud base (annual)",3250,"M VND","From original data",FMT_MN),
        ("cloud_per_1M_vcr","Cloud cost per M VCR",8,"M VND","v3: was 5",FMT_INT),
        ("vcr_per_enterprise","VCR per enterprise/month",10000,"records","Traffic estimate",FMT_INT),
        ("opex_external","External affairs (annual)",400,"M VND","Current data",FMT_MN),
        ("opex_legal","Legal & compliance (annual)",240,"M VND","Current data",FMT_MN),
        ("opex_branding","Branding & events (annual)",4000,"M VND","Current data",FMT_MN),
        ("opex_security","Security/pentest (annual)",250,"M VND","Current data",FMT_MN),
        ("opex_pilot_eval","Pilot evaluation (Y1 only)",200,"M VND","One-time",FMT_MN),
        ("payment_proc_pct","Payment gateway fee",0.025,"%","COGS on tx revenue",FMT_PCT),
        ("opex_support_per_ent","Support cost per enterprise/yr",2,"M VND","Scales w/ count",FMT_INT),
        ("opex_partner_mgmt","Partner management (annual)",500,"M VND","Relationship mgmt",FMT_MN),
        ("opex_contingency_pct","Contingency (% total OPEX)",0.10,"%","Buffer",FMT_PCT),
    ]),
    ("HR PARAMETERS",[
        ("salary_mgmt","Avg monthly — Management",45,"M VND","Director/PM",FMT_INT),
        ("salary_dev","Avg monthly — Development",35,"M VND","Engineers",FMT_INT),
        ("salary_bizops","Avg monthly — BizOps",20,"M VND","BD/support",FMT_INT),
        ("salary_data","Avg monthly — Data/Analytics",40,"M VND","Data eng",FMT_INT),
        ("salary_escalation","Annual salary increase",0.10,"%","VN market",FMT_PCT),
        ("bhxh_rate","Social insurance (employer)",0.215,"%","BHXH+BHYT+BHTN",FMT_PCT),
        ("recruit_cost_pct","Recruitment fee (% Y1 sal)",0.20,"%","Headhunter",FMT_PCT),
    ]),
    ("FINANCIAL PARAMETERS",[
        ("cit_rate","Corporate income tax",0.20,"%","Vietnamese law",FMT_PCT),
        ("discount_rate","Discount rate (WACC)",0.15,"%","v3: was 0.12",FMT_PCT),
        ("depreciation_years","Depreciation period",5,"Years","Straight-line",FMT_INT),
        ("ar_days","Accounts receivable days",45,"Days","Collection period",FMT_INT),
        ("ap_days","Accounts payable days",30,"Days","Payment terms",FMT_INT),
        ("stage2_gate","Stage 2 gate (min enterprises)",300,"","Activate tx fees",FMT_INT),
    ]),
]

for sec_title, params in sections:
    c=ws.cell(row=r,column=1,value=sec_title);c.font=SC_FONT;c.fill=SC_FILL;c.border=BD
    for col in [2,3,4]:ws.cell(row=r,column=col).fill=SC_FILL;ws.cell(row=r,column=col).border=BD
    r+=1
    c=ws.cell(row=r,column=1,value="Parameter");c.font=HD_FONT;c.fill=HD_FILL;c.alignment=Alignment(horizontal='center');c.border=BD
    for ci,h in [(2,"Value"),(3,"Unit"),(4,"Notes")]:
        c=ws.cell(row=r,column=ci,value=h);c.font=HD_FONT;c.fill=HD_FILL;c.alignment=Alignment(horizontal='center');c.border=BD
    r+=1
    for nm,lbl,val,unit,note,fmt in params:
        ws.cell(row=r,column=1,value=lbl).font=LB;ws.cell(row=r,column=1).border=BD
        c=ws.cell(row=r,column=2,value=val);c.font=PH_FONT;c.fill=PH_FILL;c.border=BD;c.number_format=fmt
        ws.cell(row=r,column=3,value=unit).font=LB;ws.cell(row=r,column=3).border=BD
        ws.cell(row=r,column=4,value=note).font=NOTE;ws.cell(row=r,column=4).border=BD
        add_nm(nm,r)
        r+=1
    r+=1

# Save
state={'named':named,'out':OUT,'final':FINAL}
with open('/tmp/v4_state.json','w') as f:json.dump(state,f)
wb.save(OUT)
print(f"v4 Part 1 done: {OUT}")
print(f"Named ranges: {len(named)}")
# Verify all named ranges
for nm,row in sorted(named.items()):
    print(f"  {nm} -> Assumptions!$B${row}")
