"""
KK2-ONLY: Cross-reference KK2. Series against VOD_NameSearch_Report_Series.csv.
Adds/updates the last column with check results. Modifies source file directly.
"""
import openpyxl
import csv
import re
import os
import shutil
from datetime import datetime

BASE = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\kiemKeK+"
XLSX = os.path.join(BASE, "List 2. Kiểm kê License_Báo cáo kiểm kê.xlsx")
CSV_SERIES = os.path.join(BASE, "VOD_NameSearch_Report_Series.csv")

# ─── Parse CSV ─────────────────────────────────────────────────────────────────
print("Parsing Series CSV...")
search_names = set()
clip_codes = set()

with open(CSV_SERIES, "r", encoding="utf-8-sig") as f:
    raw = f.read()

lines = raw.strip().split("\n")
first_data = lines[1] if len(lines) > 1 else ""
delim = ";" if first_data.count(";") > first_data.count(",") else ","

reader = csv.reader(lines, delimiter=delim)
header = next(reader)

for row in reader:
    if len(row) < 2:
        continue
    name = row[0].strip().strip('"')
    code = row[1].strip().strip('"') if len(row) > 1 else ""
    if name:
        search_names.add(name.lower().strip())
    if code and code != "Not Found":
        clip_codes.add(code.strip().upper())

print(f"  → {len(search_names)} unique search names, {len(clip_codes)} clip codes")

# ─── Extract clip codes from multi-line cell ───────────────────────────────────
def extract_clips(cell_value):
    if not cell_value:
        return []
    pattern = r'[A-Z]{2,4}_[A-Z]{2,4}(?:_[A-Z0-9]*)*_*\d{4,}'
    return list(set(re.findall(pattern, str(cell_value).upper())))

# ─── Process KK2. Series ──────────────────────────────────────────────────────
print(f"\nLoading workbook (this may take a few minutes)...")
wb = openpyxl.load_workbook(XLSX)
ws = wb["KK2. Series"]
print(f"  → KK2: {ws.max_row} rows, {ws.max_column} cols")

HEADER_ROW = 5
FIRST_DATA = 7
NEW_COL = ws.max_column + 1

# Check if a "Kết quả kiểm tra VOD" column already exists
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=HEADER_ROW, column=c).value
    if val and "kiểm tra" in str(val).lower():
        NEW_COL = c  # overwrite existing column
        print(f"  → Found existing check column at C{c}, will overwrite")
        break

ws.cell(row=HEADER_ROW, column=NEW_COL, value="Kết quả kiểm tra VOD")

# Find last data row
last_data = FIRST_DATA
for r in range(ws.max_row, FIRST_DATA - 1, -1):
    c1 = ws.cell(row=r, column=1).value
    if c1 is not None and (isinstance(c1, (int, float)) or
       (isinstance(c1, str) and c1.strip().isdigit())):
        last_data = r
        break

found = partial = not_found = 0
print(f"  → Processing rows {FIRST_DATA}-{last_data}...")

for row in range(FIRST_DATA, last_data + 1):
    c1 = ws.cell(row=row, column=1).value
    if c1 is None:
        continue

    viet_title = str(ws.cell(row=row, column=4).value or "").strip()  # C4
    eng_title = str(ws.cell(row=row, column=3).value or "").strip()   # C3
    clip_cell = ws.cell(row=row, column=2).value                       # C2
    row_clips = extract_clips(clip_cell)

    title_found = False
    if viet_title and viet_title.lower() in search_names:
        title_found = True
    if not title_found and eng_title and eng_title.lower() in search_names:
        title_found = True

    clips_found = sum(1 for c in row_clips if c in clip_codes)
    clips_total = len(row_clips)

    if title_found or (clips_total > 0 and clips_found == clips_total):
        result = "✓ Found"
        found += 1
    elif clips_found > 0:
        result = f"⚠ Partial ({clips_found}/{clips_total} clips)"
        partial += 1
    else:
        result = "✗ Not Found"
        not_found += 1

    ws.cell(row=row, column=NEW_COL, value=result)

print(f"\n  RESULTS: ✓ {found} | ⚠ {partial} | ✗ {not_found}")

# ─── Save ──────────────────────────────────────────────────────────────────────
print(f"\nSaving...")
wb.save(XLSX)
print(f"✓ Done! Column {NEW_COL} updated in KK2. Series")
