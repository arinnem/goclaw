"""
Cross-reference license inventory (KK2-KK6) against VOD CSV reports.
Adds a "Kết quả kiểm tra VOD" column to each sheet indicating if items exist in CSV.
Modifies the SOURCE Excel file directly.
"""
import openpyxl
import re
import os
import shutil
from datetime import datetime

BASE = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\kiemKeK+"
XLSX = os.path.join(BASE, "List 2. Kiểm kê License_Báo cáo kiểm kê.xlsx")

# CSV files (semicolon-delimited)
CSV_SERIES = os.path.join(BASE, "VOD_NameSearch_Report_Series.csv")
CSV_KIDS   = os.path.join(BASE, "VOD_NameSearch_Report_Kids.csv")
CSV_SHOW   = os.path.join(BASE, "VOD_NameSearch_Report_Show.csv")

# ─── CSV Parsing ───────────────────────────────────────────────────────────────

def parse_csv_auto(filepath):
    """
    Parse CSV with auto-detected delimiter (semicolon or comma). Returns:
      - search_names: set of normalized Search Name values
      - clip_codes: set of normalized Clip Code values
      - search_name_details: dict mapping search_name -> list of {clip_code, play_result, status}
    """
    import csv

    search_names = set()
    clip_codes = set()
    search_name_details = {}

    with open(filepath, "r", encoding="utf-8-sig") as f:
        raw = f.read()

    if not raw.strip():
        return search_names, clip_codes, search_name_details

    # Auto-detect delimiter: count semicolons vs commas in first data line
    lines = raw.strip().split("\n")
    if len(lines) < 2:
        return search_names, clip_codes, search_name_details

    first_data = lines[1]
    delimiter = ";" if first_data.count(";") > first_data.count(",") else ","

    # Re-parse with csv module using detected delimiter
    reader = csv.reader(lines, delimiter=delimiter)
    header = next(reader)  # skip header

    for row in reader:
        if len(row) < 2:
            continue

        search_name = row[0].strip().strip('"')
        clip_code = row[1].strip().strip('"') if len(row) > 1 else ""
        play_result = row[13].strip().strip('"') if len(row) > 13 else ""
        status = row[14].strip().strip('"') if len(row) > 14 else ""

        if search_name:
            norm_name = search_name.lower().strip()
            search_names.add(norm_name)
            if norm_name not in search_name_details:
                search_name_details[norm_name] = []
            search_name_details[norm_name].append({
                "clip_code": clip_code,
                "play_result": play_result,
                "status": status,
            })

        if clip_code and clip_code != "Not Found":
            clip_codes.add(clip_code.strip().upper())

    return search_names, clip_codes, search_name_details


def extract_clip_codes_from_cell(cell_value):
    """
    Extract individual clip codes from a multi-line cell value.
    Handles formats like:
      - EP01 – SER_DRA_2656765
      - SER_DRA_21_2652587\tLOVE IS TRUE, EP40 (VOD)
      - Ep 1: KID_CAR_21__2660472
      - Ep 2- Ep26: KID_CAR_21_2660473 - KID_CAR_21_2660497
      - ENT_CIN_HNMK__14_VN_27805\tMOM'S KITCHEN S1...
    """
    if not cell_value:
        return []

    text = str(cell_value)
    # Pattern to match clip codes: prefix_CODE_number format
    # Covers: SER_DRA_2656765, KID_CAR_21__2660472, ENT_CIN_HNMK__14_VN_27805, etc.
    pattern = r'[A-Z]{2,4}_[A-Z]{2,4}(?:_[A-Z0-9]*)*_*\d{4,}'
    codes = re.findall(pattern, text.upper())
    return list(set(codes))


# ─── Sheet Configuration ──────────────────────────────────────────────────────

SHEET_CONFIG = {
    "KK2. Series": {
        "header_row": 5,
        "first_data_row": 7,
        "viet_title_col": 4,   # C4 = VIET Title
        "eng_title_col": 3,    # C3 = ENGLISH Title
        "clip_code_col": 2,    # C2 = Clip code
        "csv_file": CSV_SERIES,
    },
    "KK3. Kids": {
        "header_row": 6,
        "first_data_row": 7,
        "viet_title_col": 4,   # C4 = VIET Title
        "eng_title_col": 3,    # C3 = ENGLISH Title
        "clip_code_col": 2,    # C2 = Clip Code
        "csv_file": CSV_KIDS,
    },
    "KK4. Show": {
        "header_row": 6,
        "first_data_row": 7,
        "viet_title_col": 4,   # C4 = Tên tiếng Việt
        "eng_title_col": 3,    # C3 = Tên tiếng Anh
        "clip_code_col": 2,    # C2 = Clip code
        "csv_file": CSV_SHOW,
    },
    "KK5. Doc": {
        "header_row": 6,
        "first_data_row": 7,
        "viet_title_col": 3,   # C3 = VIET Title
        "eng_title_col": 2,    # C2 = ENGLISH Title
        "clip_code_col": None, # No clip codes in this sheet
        "csv_file": None,      # No CSV for documentaries
    },
    "KK6. Bổ sung": {
        "header_row": 6,
        "first_data_row": 7,
        "viet_title_col": None, # No Vietnamese title
        "eng_title_col": 5,     # C5 = Program name
        "clip_code_col": 13,    # C13 has some clip codes
        "csv_file": None,       # Will search across ALL CSVs
    },
}


# ─── Main Logic ───────────────────────────────────────────────────────────────

def main():
    # Create backup
    backup_path = XLSX + f".backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    shutil.copy2(XLSX, backup_path)
    print(f"✓ Backup created: {os.path.basename(backup_path)}")

    # Load all CSVs
    csv_data = {}
    for label, path in [("Series", CSV_SERIES), ("Kids", CSV_KIDS), ("Show", CSV_SHOW)]:
        names, codes, details = parse_csv_auto(path)
        csv_data[label] = {"names": names, "codes": codes, "details": details}
        print(f"✓ Loaded {label} CSV: {len(names)} unique names, {len(codes)} clip codes")

    # Build combined lookup for KK6 (search across all CSVs)
    all_names = set()
    all_codes = set()
    all_details = {}
    for label, data in csv_data.items():
        all_names |= data["names"]
        all_codes |= data["codes"]
        for k, v in data["details"].items():
            all_details.setdefault(k, []).extend(v)

    # Load workbook (preserve formatting)
    print(f"\nLoading workbook...")
    wb = openpyxl.load_workbook(XLSX)
    print(f"✓ Loaded: {wb.sheetnames}")

    results_summary = {}

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠ Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        header_row = config["header_row"]
        first_data_row = config["first_data_row"]
        new_col = ws.max_column + 1

        # Determine which CSV data to use
        if sheet_name == "KK6. Bổ sung":
            search_names = all_names
            clip_codes_set = all_codes
        elif config["csv_file"]:
            label_map = {CSV_SERIES: "Series", CSV_KIDS: "Kids", CSV_SHOW: "Show"}
            label = label_map[config["csv_file"]]
            search_names = csv_data[label]["names"]
            clip_codes_set = csv_data[label]["codes"]
        else:
            search_names = set()
            clip_codes_set = set()

        # Write header
        ws.cell(row=header_row, column=new_col, value="Kết quả kiểm tra VOD")

        found_count = 0
        not_found_count = 0
        partial_count = 0
        no_csv_count = 0
        total_rows = 0

        # Find last data row (stop at empty rows or footer)
        last_data_row = ws.max_row
        # Walk backwards to find the last row with a number in column 1
        for r in range(ws.max_row, first_data_row - 1, -1):
            c1 = ws.cell(row=r, column=1).value
            if c1 is not None and (isinstance(c1, (int, float)) or
                (isinstance(c1, str) and (c1.strip().isdigit() or c1.startswith("K+")))):
                last_data_row = r
                break

        print(f"\n{'='*50}")
        print(f"Processing: {sheet_name} (rows {first_data_row}-{last_data_row}, new col={new_col})")

        for row in range(first_data_row, last_data_row + 1):
            # Skip empty rows
            c1 = ws.cell(row=row, column=1).value
            if c1 is None:
                continue

            total_rows += 1

            # No CSV case (KK5)
            if config["csv_file"] is None and sheet_name == "KK5. Doc":
                ws.cell(row=row, column=new_col, value="No CSV data")
                no_csv_count += 1
                continue

            # Get titles for matching
            viet_title = ""
            eng_title = ""
            if config.get("viet_title_col"):
                viet_title = str(ws.cell(row=row, column=config["viet_title_col"]).value or "").strip()
            if config.get("eng_title_col"):
                eng_title = str(ws.cell(row=row, column=config["eng_title_col"]).value or "").strip()

            # Get clip codes
            clip_codes_in_row = []
            if config.get("clip_code_col"):
                clip_cell = ws.cell(row=row, column=config["clip_code_col"]).value
                clip_codes_in_row = extract_clip_codes_from_cell(clip_cell)

            # ─── Matching Logic ───
            title_found = False
            clips_found = 0
            clips_total = len(clip_codes_in_row)

            # Check Vietnamese title match
            if viet_title:
                norm_viet = viet_title.lower().strip()
                if norm_viet in search_names:
                    title_found = True

            # Check English title match
            if not title_found and eng_title:
                norm_eng = eng_title.lower().strip()
                if norm_eng in search_names:
                    title_found = True

            # Check clip code matches
            if clip_codes_in_row:
                for code in clip_codes_in_row:
                    if code.upper() in clip_codes_set:
                        clips_found += 1

            # Determine result
            if title_found or (clips_total > 0 and clips_found == clips_total):
                result = "✓ Found"
                found_count += 1
            elif clips_found > 0:
                result = f"⚠ Partial ({clips_found}/{clips_total} clips)"
                partial_count += 1
            elif clips_total > 0 and clips_found == 0:
                result = "✗ Not Found"
                not_found_count += 1
            elif title_found:
                result = "✓ Found (title)"
                found_count += 1
            else:
                result = "✗ Not Found"
                not_found_count += 1

            ws.cell(row=row, column=new_col, value=result)

        results_summary[sheet_name] = {
            "total": total_rows,
            "found": found_count,
            "partial": partial_count,
            "not_found": not_found_count,
            "no_csv": no_csv_count,
        }

        print(f"  Total: {total_rows} | ✓ Found: {found_count} | ⚠ Partial: {partial_count} | ✗ Not Found: {not_found_count} | No CSV: {no_csv_count}")

    # Save to source file
    print(f"\nSaving to source file...")
    wb.save(XLSX)
    print(f"✓ Saved: {os.path.basename(XLSX)}")

    # Summary
    print(f"\n{'='*50}")
    print("SUMMARY")
    print(f"{'='*50}")
    for sheet, stats in results_summary.items():
        print(f"  {sheet}: {stats['found']}/{stats['total']} found, {stats['partial']} partial, {stats['not_found']} not found, {stats['no_csv']} no CSV")
    print(f"\nBackup at: {backup_path}")


if __name__ == "__main__":
    main()
