"""
Process KK3-KK6 sheets with retry save logic.
"""
import openpyxl
import csv
import re
import os
import time

BASE = r"e:\OneDrive\OneDrive - sungroup.com.vn\Works\Coding\.agent\skills\professional-writer\reports\kiemKeK+"
XLSX = os.path.join(BASE, "List 2. Kiểm kê License_Báo cáo kiểm kê.xlsx")

def parse_csv(filepath):
    search_names = set()
    clip_codes = set()
    with open(filepath, "r", encoding="utf-8-sig") as f:
        raw = f.read()
    lines = raw.strip().split("\n")
    if len(lines) < 2:
        return search_names, clip_codes
    first_data = lines[1]
    delim = ";" if first_data.count(";") > first_data.count(",") else ","
    reader = csv.reader(lines, delimiter=delim)
    next(reader)
    for row in reader:
        if len(row) < 2:
            continue
        name = row[0].strip().strip('"')
        code = row[1].strip().strip('"') if len(row) > 1 else ""
        if name:
            search_names.add(name.lower().strip())
        if code and code != "Not Found":
            clip_codes.add(code.strip().upper())
    return search_names, clip_codes

def extract_clips(cell_value):
    if not cell_value:
        return []
    pattern = r'[A-Z]{2,4}_[A-Z]{2,4}(?:_[A-Z0-9]*)*_*\d{4,}'
    return list(set(re.findall(pattern, str(cell_value).upper())))

print("Loading CSVs...")
kids_names, kids_codes = parse_csv(os.path.join(BASE, "VOD_NameSearch_Report_Kids.csv"))
show_names, show_codes = parse_csv(os.path.join(BASE, "VOD_NameSearch_Report_Show.csv"))
series_names, series_codes = parse_csv(os.path.join(BASE, "VOD_NameSearch_Report_Series.csv"))
all_names = kids_names | show_names | series_names
all_codes = kids_codes | show_codes | series_codes
print(f"  Kids: {len(kids_names)} names, {len(kids_codes)} clips")
print(f"  Show: {len(show_names)} names, {len(show_codes)} clips")
print(f"  Combined (for KK6): {len(all_names)} names, {len(all_codes)} clips")

SHEETS = {
    "KK3. Kids": {"header_row": 6, "first_data": 7, "viet_col": 4, "eng_col": 3, "clip_col": 2,
                   "names": kids_names, "codes": kids_codes},
    "KK4. Show": {"header_row": 6, "first_data": 7, "viet_col": 4, "eng_col": 3, "clip_col": 2,
                   "names": show_names, "codes": show_codes},
    "KK5. Doc":  {"header_row": 6, "first_data": 7, "viet_col": 3, "eng_col": 2, "clip_col": None,
                   "names": set(), "codes": set(), "no_csv": True},
    "KK6. Bổ sung": {"header_row": 6, "first_data": 7, "viet_col": None, "eng_col": 5, "clip_col": 13,
                      "names": all_names, "codes": all_codes},
}

print(f"\nLoading workbook...")
wb = openpyxl.load_workbook(XLSX)
print(f"  → Loaded")

for sheet_name, cfg in SHEETS.items():
    ws = wb[sheet_name]
    new_col = ws.max_column + 1
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=cfg["header_row"], column=c).value
        if val and "kiểm tra" in str(val).lower():
            new_col = c
            break
    ws.cell(row=cfg["header_row"], column=new_col, value="Kết quả kiểm tra VOD")

    last_data = cfg["first_data"]
    for r in range(ws.max_row, cfg["first_data"] - 1, -1):
        c1 = ws.cell(row=r, column=1).value
        if c1 is not None and (isinstance(c1, (int, float)) or
           (isinstance(c1, str) and (c1.strip().isdigit() or c1.startswith("K+") or c1 == "Bổ sung"))):
            last_data = r
            break

    found = partial = not_found = no_csv = 0
    for row in range(cfg["first_data"], last_data + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is None:
            continue
        if cfg.get("no_csv"):
            ws.cell(row=row, column=new_col, value="No CSV data")
            no_csv += 1
            continue

        viet = str(ws.cell(row=row, column=cfg["viet_col"]).value or "").strip() if cfg.get("viet_col") else ""
        eng = str(ws.cell(row=row, column=cfg["eng_col"]).value or "").strip() if cfg.get("eng_col") else ""
        row_clips = extract_clips(ws.cell(row=row, column=cfg["clip_col"]).value) if cfg.get("clip_col") else []

        title_found = (viet and viet.lower() in cfg["names"]) or (eng and eng.lower() in cfg["names"])
        clips_found = sum(1 for c in row_clips if c in cfg["codes"])
        clips_total = len(row_clips)

        if title_found or (clips_total > 0 and clips_found == clips_total):
            result = "✓ Found"; found += 1
        elif clips_found > 0:
            result = f"⚠ Partial ({clips_found}/{clips_total} clips)"; partial += 1
        else:
            result = "✗ Not Found"; not_found += 1

        ws.cell(row=row, column=new_col, value=result)

    print(f"  {sheet_name}: ✓ {found} | ⚠ {partial} | ✗ {not_found} | No CSV: {no_csv}")

# Retry save with backoff
for attempt in range(5):
    try:
        print(f"\nSaving (attempt {attempt+1})...")
        wb.save(XLSX)
        print(f"✓ Done!")
        break
    except PermissionError:
        print(f"  File locked, waiting 5s...")
        time.sleep(5)
else:
    # Save to alternative path
    alt = XLSX.replace(".xlsx", "_KK3_KK6_updated.xlsx")
    wb.save(alt)
    print(f"✓ Saved to alternative: {os.path.basename(alt)}")
